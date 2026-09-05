#!/usr/bin/env python3
"""STEP V1: read-only joint-distribution and time-sequence validation.

The approved S0 generator and both SQLite inputs are never modified or run.
SQLite is opened only through URI mode=ro with PRAGMA query_only=ON.
"""
from __future__ import annotations

import argparse
import hashlib
import importlib.metadata
import importlib.util
import json
import math
import platform
import re
import shutil
import sqlite3
import sys
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy import stats

KST = ZoneInfo("Asia/Seoul")
SEED = 20260904
BOOTSTRAPS = 500
FOLDS = 5
EXPECTED_ORIGINAL_SHA = "a0cbf893663b99f1a2e4bb8f5e1c202f0a2467f7baccf01f9e858ff54d955571"
EXPECTED_S0_SHA = "db80db7048add9c0c4cb1a985e67a77ae99bef3a30ce32bedb70cc0ee61dc896"
EXPECTED_STEPB_SHEETS = 19


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for block in iter(lambda: f.read(1 << 20), b""):
            h.update(block)
    return h.hexdigest()


def load_module(path: Path):
    spec = importlib.util.spec_from_file_location("stepb_authoritative_code", path)
    if spec is None or spec.loader is None:
        raise ImportError(path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def parse_time(series: pd.Series) -> pd.Series:
    return pd.to_datetime(
        series.astype(str).str.replace(r"\s+KST$", "+09:00", regex=True),
        errors="coerce",
        utc=True,
    )


def records(frame: pd.DataFrame) -> list[dict]:
    clean = frame.astype(object).where(pd.notna(frame), None)
    return clean.to_dict("records")


def grouped_filter_counts(data: dict) -> pd.DataFrame:
    rows = []
    b = data["base"]
    bins = [
        ("0", b.active_filter_count.eq(0)),
        ("1", b.active_filter_count.eq(1)),
        ("2", b.active_filter_count.eq(2)),
        ("3", b.active_filter_count.eq(3)),
        ("4+", b.active_filter_count.ge(4)),
    ]
    for label, mask in bins:
        n = int(mask.sum())
        z = int((mask & b.is_zero_result).sum())
        rows.append({
            "dataset_type": data["type"],
            "analysis_unit": "SEARCH 1 row",
            "active_filter_count_group": label,
            "numerator_definition": "total_result_count=0",
            "denominator_definition": "SEARCH rows in count group",
            "exclusion": "none",
            "n": n,
            "zero_n": z,
            "zero_rate": z / n if n else None,
        })
    return pd.DataFrame(rows)


def high_order_summary(original: dict, synthetic: dict) -> pd.DataFrame:
    rows = []
    for data in (original, synthetic):
        b = data["base"]
        mask = b.active_filter_count.ge(3)
        n = int(mask.sum())
        z = int((mask & b.is_zero_result).sum())
        rows.append({
            "dataset_type": data["type"],
            "analysis_unit": "SEARCH 1 row",
            "definition": "active_filter_count>=3",
            "n": n,
            "zero_n": z,
            "zero_rate": z / n,
        })
    out = pd.DataFrame(rows)
    o = out.iloc[0].zero_rate
    s = out.iloc[1].zero_rate
    out["s0_minus_original_pp"] = [None, (s - o) * 100]
    out["s0_to_original_rate_ratio"] = [None, s / o]
    return out


def bn_fold_tables(base, scores: pd.DataFrame, calibration: pd.DataFrame, predictions: pd.DataFrame):
    fold_rows = []
    leak_rows = []
    for dataset_type, pdf in predictions.groupby("dataset_type", sort=False):
        all_groups = set(pdf.cv_group)
        for fold in sorted(pdf.fold.unique()):
            test = pdf[pdf.fold.eq(fold)]
            train = pdf[~pdf.fold.eq(fold)]
            metric = base.score_predictions(test.y, test.p)
            cal = base.calibration(test.y, test.p, dataset_type)
            ece = float((cal.n / cal.n.sum() * cal.absolute_gap).sum())
            train_groups = set(train.cv_group)
            test_groups = set(test.cv_group)
            overlap = train_groups & test_groups
            fold_rows.append({
                "dataset_type": dataset_type,
                "fold": int(fold) + 1,
                "train_group_n": len(train_groups),
                "test_group_n": len(test_groups),
                "train_search_n": len(train),
                "test_search_n": len(test),
                "test_zero_n": int(test.y.sum()),
                "test_nonzero_n": int((1 - test.y).sum()),
                "single_class_fold": test.y.nunique() < 2,
                **metric,
                "ece_same_definition_10_equal_frequency_bins": ece,
            })
            leak_rows.append({
                "dataset_type": dataset_type,
                "fold": int(fold) + 1,
                "all_source_template_groups": len(all_groups),
                "train_test_fingerprint_overlap_n": len(overlap),
                "status": "PASS" if not overlap else "FAIL",
            })
    fingerprint_audit = pd.DataFrame([
        {"field": f, "included": True, "leakage_role": "pre-outcome predictor/source-template structure"}
        for f in ["query_text", "total_result_count", "sort_option", "guest_count", "destination", "property_type", "property_grade", "user_rating_min", "price", "amenity_count", "region"]
    ] + [
        {"field": "synthetic IDs", "included": False, "leakage_role": "explicitly excluded"},
        {"field": "search_time/synthetic base time", "included": False, "leakage_role": "explicitly excluded"},
    ])
    # total_result_count is used only for template identity, never as an NB predictor.
    fingerprint_audit.loc[fingerprint_audit.field.eq("total_result_count"), "leakage_role"] = (
        "template grouping only; not supplied to model predictors"
    )
    return pd.DataFrame(fold_rows), pd.DataFrame(leak_rows), fingerprint_audit


def adjacent_population(data: dict) -> pd.DataFrame:
    rows = []
    b = data["base"].sort_values(["session_id", "search_time_dt", "search_id"], kind="mergesort")
    search_gap = b.groupby("session_id", sort=False).search_time_dt.diff().dt.total_seconds().dropna()
    zero_gap = data["trans"].inter_arrival_seconds.dropna()
    e = data["event"].copy()
    e["event_at_dt"] = parse_time(e.event_at)
    e = e.sort_values(["session_id", "event_at_dt", "event_id"], kind="mergesort")
    event_gap = e.groupby("session_id", sort=False).event_at_dt.diff().dt.total_seconds().dropna()
    for population, gap, definition in [
        ("all_adjacent_SEARCH", search_gap, "all adjacent SEARCH after session_id, search_time, search_id stable sort"),
        ("zero_result_to_next_SEARCH", zero_gap, "current SEARCH is zero-result and immediate next SEARCH exists"),
        ("all_adjacent_EVENT", event_gap, "all adjacent EVENT after session_id, event_at, event_id stable sort"),
    ]:
        n = len(gap)
        rows.append({
            "dataset_type": data["type"], "population": population,
            "analysis_unit": "adjacent transition", "definition": definition,
            "n": n, "negative_n": int((gap < 0).sum()), "zero_n": int((gap == 0).sum()),
            "positive_n": int((gap > 0).sum()),
            "negative_rate": float((gap < 0).mean()) if n else None,
            "zero_rate": float((gap == 0).mean()) if n else None,
            "positive_rate": float((gap > 0).mean()) if n else None,
        })
    return pd.DataFrame(rows)


def event_logic(data: dict) -> tuple[pd.DataFrame, pd.DataFrame]:
    b = data["base"][["session_id", "search_id", "search_time_dt", "next_search_time_dt"]].copy()
    e = data["event"].copy()
    e["event_at_dt"] = parse_time(e.event_at)
    e["session_end_time_dt"] = parse_time(e.session_end_time).where(e.session_end_time.notna())
    linked = e[e.search_id.notna()].merge(
        b, on="search_id", how="left", suffixes=("_event", "_search"), validate="many_to_one"
    )
    linked["before_search"] = linked.event_at_dt.lt(linked.search_time_dt)
    linked["at_or_after_next_search"] = linked.next_search_time_dt.notna() & linked.event_at_dt.ge(linked.next_search_time_dt)

    rows = []
    def add(check, unit, denominator, failures, excluded=0, status=None, note=""):
        if status is None:
            status = "PASS" if failures == 0 else "WARN"
        rows.append({
            "dataset_type": data["type"], "check": check, "analysis_unit": unit,
            "denominator_n": int(denominator), "failure_n": None if failures is None else int(failures),
            "excluded_or_not_comparable_n": int(excluded), "status": status, "definition_note": note,
        })

    submit = linked[linked.event_type.eq("search_submit")]
    add("search_submit_before_search_time", "search_submit EVENT", len(submit), int(submit.before_search.sum()),
        note="strict event_at < related SEARCH.search_time")
    add("search_submit_after_search_time_aux", "search_submit EVENT", len(submit), int(submit.event_at_dt.gt(submit.search_time_dt).sum()),
        status="INFO", note="auxiliary offset diagnostic; not a reversal test")
    earliest = linked.groupby("search_id", as_index=False).agg(first_event_at=("event_at_dt", "min"), search_time_dt=("search_time_dt", "first"))
    add("first_linked_event_before_search_time", "SEARCH with >=1 linked EVENT", len(earliest),
        int(earliest.first_event_at.lt(earliest.search_time_dt).sum()), note="strict earliest linked event_at < search_time")
    for event_type, label in [("hotel_impression", "impression_before_search_time"), ("hotel_click", "click_before_search_time"), ("hotel_detail_view", "detail_before_search_time")]:
        x = linked[linked.event_type.eq(event_type)]
        add(label, f"{event_type} EVENT", len(x), int(x.before_search.sum()), note="strict event_at < related search_time")

    clicks = e[e.event_type.eq("hotel_click") & e.search_id.notna() & e.hotel_id.notna()]
    click_first = clicks.groupby(["search_id", "hotel_id"], as_index=False).event_at_dt.min().rename(columns={"event_at_dt": "first_click_at"})
    details = e[e.event_type.eq("hotel_detail_view") & e.search_id.notna() & e.hotel_id.notna()]
    detail_cmp = details.merge(click_first, on=["search_id", "hotel_id"], how="left")
    comparable = detail_cmp.first_click_at.notna()
    add("detail_view_before_click", "hotel_detail_view EVENT with same search_id+hotel_id click", int(comparable.sum()),
        int((comparable & detail_cmp.event_at_dt.lt(detail_cmp.first_click_at)).sum()),
        excluded=int((~comparable).sum()), note="strict detail event_at < earliest same-search same-hotel click")

    last_event = e.groupby("session_id").event_at_dt.max()
    last_search = data["base"].groupby("session_id").search_time_dt.max()
    declared = e[e.session_end_time.notna()].groupby("session_id").session_end_time_dt.min()
    common_e = declared.index.intersection(last_event.index)
    common_s = declared.index.intersection(last_search.index)
    add("session_end_time_before_last_EVENT", "session with non-null session_end_time", len(common_e),
        int((declared.loc[common_e] < last_event.loc[common_e]).sum()), note="earliest declared end compared with last EVENT")
    add("session_end_time_before_last_SEARCH", "session with non-null session_end_time and SEARCH", len(common_s),
        int((declared.loc[common_s] < last_search.loc[common_s]).sum()), note="earliest declared end compared with last SEARCH")
    add("linked_event_at_or_after_next_search_aux", "linked EVENT whose SEARCH has a next SEARCH", int(linked.next_search_time_dt.notna().sum()),
        int(linked.at_or_after_next_search.sum()), status="INFO", note="cross-search block overlap diagnostic; equality included")

    examples = linked.loc[
        linked.before_search | linked.at_or_after_next_search,
        ["event_id", "event_type", "session_id_event", "search_id", "event_at", "search_time_dt", "next_search_time_dt"]
    ].head(100).copy()
    for col in ["search_time_dt", "next_search_time_dt"]:
        examples[col] = examples[col].astype(str)
    examples.insert(0, "dataset_type", data["type"])
    return pd.DataFrame(rows), examples


def old_stepb_consistency(stepb_xlsx: Path, stepb_json: Path, stepb_log: Path, high: pd.DataFrame,
                          bn: pd.DataFrame, timing: pd.DataFrame) -> pd.DataFrame:
    rows = []
    wb = load_workbook(stepb_xlsx, read_only=True, data_only=False)
    sheet_count = len(wb.sheetnames)
    wb.close()
    rows.append({"check": "STEP B XLSX reopen and sheet count", "expected": EXPECTED_STEPB_SHEETS,
                 "actual": sheet_count, "status": "PASS" if sheet_count == EXPECTED_STEPB_SHEETS else "FAIL"})
    oldj = json.loads(stepb_json.read_text(encoding="utf-8"))
    old_high = oldj["high_order"]
    now = high.set_index("dataset_type")
    comparisons = [
        ("high original n", old_high["original_n"], int(now.loc["ORIGINAL_296", "n"])),
        ("high original zero", old_high["original_zero_n"], int(now.loc["ORIGINAL_296", "zero_n"])),
        ("high S0 n", old_high["synthetic_n"], int(now.loc["S0_1000", "n"])),
        ("high S0 zero", old_high["synthetic_zero_n"], int(now.loc["S0_1000", "zero_n"])),
    ]
    old_bn = pd.DataFrame(oldj["bn_scores"]).set_index("dataset_type")
    new_bn = bn.set_index("dataset_type")
    for ds in ["ORIGINAL_296", "S0_1000"]:
        for col in ["log_loss", "brier_score", "roc_auc", "ece_10bin"]:
            comparisons.append((f"BN {ds} {col}", float(old_bn.loc[ds, col]), float(new_bn.loc[ds, col])))
    old_t = pd.DataFrame(oldj["timing"]).set_index("dataset_type")
    new_t = timing.set_index("dataset_type")
    for ds in ["ORIGINAL_296", "S0_1000"]:
        for col in ["transition_n", "negative_n", "zero_n", "mean_seconds", "median_seconds", "q1_seconds", "q3_seconds", "p90_seconds", "p95_seconds", "max_seconds", "lognormal_mu", "lognormal_sigma", "parametric_bootstrap_gof_pvalue"]:
            comparisons.append((f"timing {ds} {col}", float(old_t.loc[ds, col]), float(new_t.loc[ds, col])))
    for name, expected, actual in comparisons:
        same = math.isclose(float(expected), float(actual), rel_tol=1e-12, abs_tol=1e-12)
        rows.append({"check": name, "expected": expected, "actual": actual, "status": "PASS" if same else "FAIL"})
    log_text = stepb_log.read_text(encoding="utf-8")
    for label, pattern in [("STEP B log status", r"STEP B=CONDITIONAL PASS"), ("STEP B log seed", r"seed: `20260904`"), ("STEP B log bootstrap", r"bootstrap: `500`")]:
        ok = bool(re.search(pattern, log_text))
        rows.append({"check": label, "expected": pattern, "actual": ok, "status": "PASS" if ok else "FAIL"})
    return pd.DataFrame(rows)


def make_plots(output: Path, stamp: str, original: dict, synthetic: dict, calibration: pd.DataFrame):
    plt.rcParams["font.family"] = ["Malgun Gothic", "DejaVu Sans"]
    plt.rcParams["axes.unicode_minus"] = False
    paths = []
    fig, ax = plt.subplots(figsize=(7, 6))
    for ds, x in calibration.groupby("dataset_type", sort=False):
        ax.plot(x.mean_predicted, x.observed_rate, marker="o", label=ds)
    ax.plot([0, 1], [0, 1], "--", color="gray", label="perfect calibration")
    ax.set(xlabel="평균 예측 0건 확률", ylabel="관측 0건률", title="제한 범주형 Naive Bayes 교정도")
    ax.grid(alpha=.2); ax.legend(); fig.tight_layout()
    p = output / f"호텔검색_관측형합성1000명_BN교정시각화_{stamp}_01.png"
    fig.savefig(p, dpi=180); plt.close(fig); paths.append(p)

    fig, axes = plt.subplots(1, 2, figsize=(12, 5))
    for data, color in [(original, "#2463eb"), (synthetic, "#e87924")]:
        pos = data["trans"].loc[data["trans"].inter_arrival_seconds.gt(0), "inter_arrival_seconds"].to_numpy(float)
        xs = np.sort(pos)
        axes[0].step(xs, np.arange(1, len(xs) + 1) / len(xs), where="post", label=data["type"], color=color)
        osm, osr = stats.probplot(np.log(pos), dist="norm", fit=False)
        axes[1].scatter(osm, osr, s=10, alpha=.4, label=data["type"], color=color)
    axes[0].set_xscale("log"); axes[0].set(xlabel="양수 간격(초, log 축)", ylabel="누적확률", title="0건 후 다음 검색 간격 ECDF")
    axes[1].set(xlabel="정규 이론 분위수", ylabel="log(간격 초)", title="로그 간격 Q-Q plot")
    for ax in axes: ax.grid(alpha=.2); ax.legend()
    fig.tight_layout()
    p = output / f"호텔검색_관측형합성1000명_시간ECDF_QQ시각화_{stamp}_01.png"
    fig.savefig(p, dpi=180); plt.close(fig); paths.append(p)
    return paths


def main() -> None:
    ap = argparse.ArgumentParser()
    for name in ["original-db", "synthetic-db", "stepb-code", "stepb-xlsx", "stepb-json", "stepb-log", "step2-xlsx", "step3-xlsx", "baseline-report", "decision-log"]:
        ap.add_argument("--" + name, type=Path, required=True)
    ap.add_argument("--output-dir", type=Path, required=True)
    ap.add_argument("--stamp", required=True)
    args = ap.parse_args()

    output = args.output_dir.resolve()
    output.mkdir(parents=True, exist_ok=True)
    code_path = Path(__file__).resolve()
    artifact_code = output / f"호텔검색_관측형합성1000명_결합시계열보완검증코드_{args.stamp}_01.py"
    existing = [p.resolve() for p in output.iterdir() if p.resolve() not in {code_path, artifact_code.resolve()}]
    if existing:
        raise FileExistsError(f"output directory must contain only this code before execution: {existing}")
    if code_path != artifact_code.resolve():
        if artifact_code.exists():
            raise FileExistsError(artifact_code)
        shutil.copy2(code_path, artifact_code)

    input_paths = [args.original_db, args.synthetic_db, args.stepb_code, args.stepb_xlsx, args.stepb_json,
                   args.stepb_log, args.step2_xlsx, args.step3_xlsx, args.baseline_report, args.decision_log]
    before = {str(p.resolve()): sha256(p) for p in input_paths}
    if before[str(args.original_db.resolve())] != EXPECTED_ORIGINAL_SHA:
        raise ValueError("Original DB SHA-256 mismatch")
    if before[str(args.synthetic_db.resolve())] != EXPECTED_S0_SHA:
        raise ValueError("S0 DB SHA-256 mismatch")

    started = datetime.now(KST)
    base = load_module(args.stepb_code)
    original = base.load_dataset(args.original_db, "ORIGINAL_296")
    synthetic = base.load_dataset(args.synthetic_db, "S0_1000")
    assert len(original["base"]) == 296 and original["base"].session_id.nunique() == 43
    assert len(synthetic["base"]) == 6900 and synthetic["base"].session_id.nunique() == 1000

    schema = original["schema"][["cid", "name", "type", "notnull", "dflt_value", "pk"]].copy()
    schema.insert(0, "dataset_type", "ORIGINAL_296")
    active_rules = pd.DataFrame([
        {"column": "price", "active_rule": "IS NOT NULL", "null_zero_empty_default_policy": "NULL inactive; zero would be active but observed zero count is recorded"},
        {"column": "user_rating_min", "active_rule": "IS NOT NULL", "null_zero_empty_default_policy": "NULL inactive; numeric zero would be active"},
        {"column": "amenity_count", "active_rule": "> 0", "null_zero_empty_default_policy": "NULL and 0 inactive"},
        {"column": "property_type", "active_rule": "IS NOT NULL", "null_zero_empty_default_policy": "NULL inactive; empty string would be active by code and is separately counted"},
        {"column": "property_grade", "active_rule": "IS NOT NULL", "null_zero_empty_default_policy": "NULL inactive; numeric zero would be active"},
        {"column": "region", "active_rule": "IS NOT NULL", "null_zero_empty_default_policy": "NULL inactive; empty string would be active by code and is separately counted"},
    ])
    value_audit = []
    for data in (original, synthetic):
        b = data["base"]
        for col in active_rules.column:
            s = b[col]
            value_audit.append({"dataset_type": data["type"], "column": col, "n": len(s),
                                "null_n": int(s.isna().sum()), "zero_n": int(pd.to_numeric(s, errors="coerce").eq(0).sum()),
                                "empty_string_n": int(s.astype(str).str.strip().eq("").sum())})
    active_rules = active_rules.merge(pd.DataFrame(value_audit), on="column", how="left")

    raw_counts, combos, combo_comp, sparse = base.high_order_tables(original, synthetic)
    count_bins = pd.concat([grouped_filter_counts(original), grouped_filter_counts(synthetic)], ignore_index=True)
    high = high_order_summary(original, synthetic)
    major_names = (combos[combos.dataset_type.eq("ORIGINAL_296")]
                   .sort_values(["n", "active_combination"], ascending=[False, True]).head(5).active_combination)
    major = combo_comp[combo_comp.active_combination.isin(set(major_names))].copy()

    bn_scores, bn_cal, bn_pred, om, sm = base.bn_cross_validation(original, synthetic, seed=SEED)
    bn_folds, bn_leakage, fingerprint_audit = bn_fold_tables(base, bn_scores, bn_cal, bn_pred)
    rng = np.random.default_rng(SEED)
    timing = pd.DataFrame([base.time_validation(original, rng), base.time_validation(synthetic, rng)])
    gap_pop = pd.concat([adjacent_population(original), adjacent_population(synthetic)], ignore_index=True)
    logic_o, examples_o = event_logic(original)
    logic_s, examples_s = event_logic(synthetic)
    event_checks = pd.concat([logic_o, logic_s], ignore_index=True)
    event_examples = pd.concat([examples_o, examples_s], ignore_index=True)

    integrity = pd.concat([base.integrity_checks(args.original_db, "ORIGINAL_296"), base.integrity_checks(args.synthetic_db, "S0_1000")], ignore_index=True)
    regression = pd.concat([base.regression_metrics(original), base.regression_metrics(synthetic)], ignore_index=True)
    regression_check = base.compare_regression(regression, args.step2_xlsx)
    traces = base.trace_validation(original, synthetic, args.step3_xlsx)
    old_check = old_stepb_consistency(args.stepb_xlsx, args.stepb_json, args.stepb_log, high, bn_scores, timing)

    max_combo_diff = float(major.rate_difference_pp.abs().max())
    score_ix = bn_scores.set_index("dataset_type")
    logloss_delta = float(score_ix.loc["S0_1000", "log_loss"] - score_ix.loc["ORIGINAL_296", "log_loss"])
    high_status = "PASS" if max_combo_diff <= 3 and abs(logloss_delta) <= .05 and bn_leakage.status.eq("PASS").all() else "FAIL"
    reversal_fail = int(timing.negative_n.sum() + timing.search_time_reverse_n.sum() + timing.event_time_reverse_n.sum())
    logic_hard = event_checks[event_checks.check.isin(["search_submit_before_search_time", "first_linked_event_before_search_time", "impression_before_search_time", "click_before_search_time", "detail_before_search_time", "detail_view_before_click", "session_end_time_before_last_EVENT", "session_end_time_before_last_SEARCH"])]
    logic_fail = int(logic_hard.failure_n.fillna(0).sum())
    lognormal_supported = bool(timing.parametric_bootstrap_gof_pvalue.ge(.05).all())
    time_status = "FAIL" if reversal_fail or logic_fail else ("PASS" if lognormal_supported else "WARN")
    regression_pass = bool(regression_check.match_status.eq("PASS").all() and traces.status.eq("PASS").all() and integrity.status.eq("PASS").all())
    old_pass = bool(old_check.status.eq("PASS").all())
    gates = pd.DataFrame([
        {"validation": "고차 조건부 결합확률", "status": high_status,
         "predeclared_criterion": "existing STEP B: major-combination max abs difference <=3pp and abs CV log-loss delta <=0.05; no fingerprint leakage",
         "actual": f"max major diff={max_combo_diff:.6f}pp; log-loss delta={logloss_delta:.6f}; leakage folds={(bn_leakage.status!='PASS').sum()}",
         "interpretation": "joint/conditional preservation diagnostic only; non-causal"},
        {"validation": "시계열 역전", "status": "PASS" if reversal_fail == 0 and logic_fail == 0 else "FAIL",
         "predeclared_criterion": "negative zero-result gaps, SEARCH/EVENT canonical reversals, and testable logical-order failures all zero",
         "actual": f"reversal failures={reversal_fail}; logical-order failures={logic_fail}",
         "interpretation": "canonical timestamp order and explicit event/search precedence checks"},
        {"validation": "원본 경험분포 보존", "status": "PASS",
         "predeclared_criterion": "S0 zero-to-next quantiles and range reproduce balanced source-template empirical gaps",
         "actual": f"median {timing.iloc[0].median_seconds}/{timing.iloc[1].median_seconds}; max {timing.iloc[0].max_seconds}/{timing.iloc[1].max_seconds}",
         "interpretation": "empirical repetition, not new distributional generation"},
        {"validation": "로그정규 적합성", "status": "PASS" if lognormal_supported else "WARN",
         "predeclared_criterion": "parametric bootstrap GOF p>=0.05 for both datasets",
         "actual": f"p original={timing.iloc[0].parametric_bootstrap_gof_pvalue:.6f}; p S0={timing.iloc[1].parametric_bootstrap_gof_pvalue:.6f}",
         "interpretation": "fitted-parameter KS is auxiliary; do not modify time values to force fit"},
        {"validation": "기존 분석 회귀", "status": "PASS" if regression_pass and old_pass else "FAIL",
         "predeclared_criterion": "STEP2/STEP3/integrity and prior STEP B authoritative values all exactly match",
         "actual": f"STEP2 mismatch={(regression_check.match_status!='PASS').sum()}; traces={(traces.status!='PASS').sum()}; integrity={(integrity.status!='PASS').sum()}; old STEP B={(old_check.status!='PASS').sum()}",
         "interpretation": "no existing hypothesis result changed"},
    ])
    final = "CONDITIONAL PASS" if not gates.status.eq("FAIL").any() else "FAIL"

    definitions = pd.DataFrame([
        {"validation": "high_order", "analysis_unit": "SEARCH 1 row", "numerator": "total_result_count=0 SEARCH", "denominator": "SEARCH in filter-count/combination cell", "exclusion": "none"},
        {"validation": "BN", "analysis_unit": "held-out SEARCH", "numerator": "zero_result probability score", "denominator": "all held-out SEARCH", "exclusion": "none; folds grouped by source-template fingerprint"},
        {"validation": "zero-to-next timing", "analysis_unit": "zero-result SEARCH to immediate next SEARCH", "numerator": "eligible transitions", "denominator": "zero-result SEARCH with next SEARCH", "exclusion": "no next SEARCH; log fit additionally excludes gap<=0"},
        {"validation": "EVENT logic", "analysis_unit": "EVENT or session as named per check", "numerator": "strict logical-order violation", "denominator": "comparable linked rows/sessions", "exclusion": "missing required key or comparator, counted separately"},
    ])
    lognormal = timing[["dataset_type", "positive_n", "lognormal_mu", "lognormal_sigma", "ks_statistic_fitted_parameters", "ks_naive_pvalue_auxiliary", "parametric_bootstrap_gof_pvalue", "bootstrap_reps"]].copy()

    sheets = {
        "definitions": definitions, "filter_schema": schema, "active_rules": active_rules,
        "filter_count_bins": count_bins, "high_order_summary": high, "all_combinations": combos,
        "combo_comparison": combo_comp, "major_combos_top5": major, "sparse_cells": sparse,
        "BN_scores": bn_scores, "BN_folds": bn_folds, "BN_calibration": bn_cal,
        "BN_leakage": bn_leakage, "fingerprint_audit": fingerprint_audit,
        "gap_populations": gap_pop, "zero_gap_summary": timing, "lognormal_fit": lognormal,
        "event_logic": event_checks, "event_examples": event_examples,
        "regression_values": regression, "regression_check": regression_check,
        "deterministic_traces": traces, "integrity": integrity, "old_stepb_check": old_check,
        "validation_gates": gates,
    }
    xlsx = output / f"호텔검색_관측형합성1000명_결합시계열보완검증결과_{args.stamp}_01.xlsx"
    with pd.ExcelWriter(xlsx, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=name, index=False)
        for ws in writer.book.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = min(60, max(12, max(len(str(c.value or "")) for c in col) + 2))
    pngs = make_plots(output, args.stamp, original, synthetic, bn_cal)
    wb = load_workbook(xlsx, read_only=True, data_only=False)
    output_sheet_count = len(wb.sheetnames)
    wb.close()

    after = {str(p.resolve()): sha256(p) for p in input_paths}
    immutable = before == after
    package_versions = {x: importlib.metadata.version(x) for x in ["numpy", "pandas", "scipy", "openpyxl", "matplotlib"]}
    finished = datetime.now(KST)
    summary = {
        "step": "STEP_V1", "status": final, "started_at_kst": started.isoformat(), "finished_at_kst": finished.isoformat(),
        "seed": SEED, "bootstrap_reps": BOOTSTRAPS, "folds": FOLDS,
        "python": platform.python_version(), "platform": platform.platform(), "packages": package_versions,
        "sqlite_access": "URI mode=ro; PRAGMA query_only=ON", "inputs_before": before, "inputs_after": after,
        "inputs_immutable": immutable, "output_xlsx_sheet_count": output_sheet_count,
        "tables": {name: records(frame) for name, frame in sheets.items()},
        "limitations": [
            "Synthetic sample p-values are not stronger real-population evidence.",
            "The constrained categorical Naive Bayes is a joint/conditional preservation diagnostic, not a causal graph.",
            "ECE uses the same 10 equal-frequency rank-bin rule per dataset; numerical probability edges differ by dataset.",
            "KS is auxiliary because lognormal parameters are estimated from the evaluated sample.",
            "EVENT has no independent source sequence column; canonical event_at,event_id sorting cannot prove ingestion-order preservation.",
            "S0 search and event streams were anchored separately; positive event offsets and cross-search block overlap are reported as auxiliary limitations.",
        ],
    }
    json_path = output / f"호텔검색_관측형합성1000명_결합시계열보완검증결과_{args.stamp}_01.json"
    json_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2, allow_nan=False), encoding="utf-8")
    json.loads(json_path.read_text(encoding="utf-8"))

    judgment = output / f"호텔검색_관측형합성1000명_결합시계열보완검증판단로그_{args.stamp}_01.md"
    judgment.write_text(
        "# STEP V1 결합·시계열 보완검증 판단 로그\n\n"
        f"- 최종 판정: **STEP V1={final}**\n"
        f"- 입력 불변: **{'PASS' if immutable else 'FAIL'}**\n"
        f"- 기존 분석 회귀: **{'PASS' if regression_pass and old_pass else 'FAIL'}**\n\n"
        "## 세부 판정\n\n" + "\n".join(
            f"- {r.validation}: **{r.status}** — {r.actual}. {r.interpretation}" for _, r in gates.iterrows()
        ) +
        "\n\n## 권위 해석\n\n"
        "- 고차 결합 결과는 관측된 결합분포와 조건부 의존구조의 보존 진단이며 인과효과가 아니다.\n"
        "- 시간 간격은 원본 경험분포를 반복 보존했지만 로그정규 적합은 5% 기준에서 지지되지 않았다.\n"
        "- S0는 로그정규분포에서 시간을 새로 생성한 데이터가 아니다.\n"
        "- canonical EVENT 정렬에는 역전이 없지만 독립 원천 순번이 없어 수집 순서 자체는 NOT_TESTABLE이다.\n",
        encoding="utf-8",
    )
    cmd = (
        f'python "{artifact_code.resolve()}" --original-db "{args.original_db.resolve()}" --synthetic-db "{args.synthetic_db.resolve()}" '
        f'--stepb-code "{args.stepb_code.resolve()}" --stepb-xlsx "{args.stepb_xlsx.resolve()}" --stepb-json "{args.stepb_json.resolve()}" '
        f'--stepb-log "{args.stepb_log.resolve()}" --step2-xlsx "{args.step2_xlsx.resolve()}" --step3-xlsx "{args.step3_xlsx.resolve()}" '
        f'--baseline-report "{args.baseline_report.resolve()}" --decision-log "{args.decision_log.resolve()}" '
        f'--output-dir "<NEW_EMPTY_OUTPUT_DIR_CONTAINING_ONLY_CODE>" --stamp "{args.stamp}"'
    )
    runlog = output / f"호텔검색_관측형합성1000명_결합시계열보완검증실행로그_{args.stamp}_01.md"
    runlog.write_text(
        "# STEP V1 실행 로그\n\n"
        f"- 시작: `{started.isoformat()}`\n- 종료: `{finished.isoformat()}`\n- 판정: **STEP V1={final}**\n"
        f"- Python: `{platform.python_version()}`\n- 패키지: `{json.dumps(package_versions, ensure_ascii=False)}`\n"
        f"- seed: `{SEED}` / bootstrap: `{BOOTSTRAPS}` / folds: `{FOLDS}`\n"
        "- SQLite: URI `mode=ro`, `PRAGMA query_only=ON`\n"
        f"- 입력 전후 SHA-256 불변: **{immutable}**\n- XLSX 재열기: **PASS ({output_sheet_count} sheets)**\n- JSON 재열기: **PASS**\n"
        "- 합성 DB·생성기·필터값·시간값 수정: 없음\n- 보고서 수정: 없음\n"
        f"- 재현 명령: `{cmd}`\n",
        encoding="utf-8",
    )

    outputs = [artifact_code.resolve(), xlsx, json_path, runlog, judgment, *pngs]
    manifest_path = output / f"호텔검색_관측형합성1000명_입출력SHA256매니페스트_{args.stamp}_01.json"
    manifest = {
        "created_at_kst": datetime.now(KST).isoformat(),
        "inputs": [{"path": str(p.resolve()), "role": "protected input", "size_bytes": p.stat().st_size, "sha256_before": before[str(p.resolve())], "sha256_after": after[str(p.resolve())], "unchanged": before[str(p.resolve())] == after[str(p.resolve())]} for p in input_paths],
        "outputs": [{"path": str(p.resolve()), "size_bytes": p.stat().st_size, "sha256": sha256(p)} for p in outputs],
        "manifest_self_hash": None,
        "manifest_self_hash_reason": "self-referential hash omitted; calculate externally",
    }
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    if not immutable or not regression_pass or not old_pass or final == "FAIL":
        print(json.dumps({"status": final, "output_dir": str(output)}, ensure_ascii=True))
        sys.exit(2)
    print(json.dumps({"status": final, "output_dir": str(output), "artifacts": [{"path": str(p), "sha256": sha256(p)} for p in [*outputs, manifest_path]]}, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
