#!/usr/bin/env python3
"""STEP R1: rebuild approved S0 with one common time origin per source session."""
from __future__ import annotations

import argparse
import hashlib
import json
import sqlite3
import sys
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from statistics import median
from zoneinfo import ZoneInfo

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook

KST = ZoneInfo("Asia/Seoul")
TABLES = ["user", "hotel", "room", "search", "search_filter", "search_result", "event", "booking"]
BOOKING_EVENTS = {"booking_start", "booking_complete", "booking_cancel"}
HOTEL_EVENTS = {"hotel_click", "hotel_detail_view"}


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for block in iter(lambda: f.read(1 << 20), b""):
            h.update(block)
    return h.hexdigest()


def parse_ts(value):
    if value is None or pd.isna(value) or str(value).strip() == "":
        return None
    return datetime.fromisoformat(str(value).replace(" KST", "+09:00"))


def format_ts(value):
    if value is None:
        return None
    return value.astimezone(KST).strftime("%Y-%m-%d %H:%M:%S KST")


def ro(path: Path):
    conn = sqlite3.connect(path.resolve().as_uri() + "?mode=ro", uri=True)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA query_only=ON")
    assert conn.execute("PRAGMA query_only").fetchone()[0] == 1
    return conn


def write_rows(conn, table, rows):
    if not rows:
        return
    cols = list(rows[0])
    conn.executemany(
        f'INSERT INTO "{table}"({",".join(cols)}) VALUES({",".join("?" for _ in cols)})',
        [[row.get(col) for col in cols] for row in rows],
    )


def shifted(value, base, origin):
    parsed = parse_ts(value)
    return None if parsed is None else format_ts(base + (parsed - origin))


def replicated_event_rows(events, results, source_session):
    result_pairs = {(r.search_id, r.hotel_id) for r in results.itertuples()}
    rows = []
    for ordinal, row in enumerate(
        events[events.session_id == source_session].sort_values(["event_at", "event_id"], na_position="last").itertuples(), 1
    ):
        if row.event_type in BOOKING_EVENTS:
            continue
        if row.event_type in HOTEL_EVENTS and (row.search_id, row.hotel_id) not in result_pairs:
            continue
        rows.append((ordinal, row))
    return rows


def build(source_path, parent_path, output_path, cfg, code_hash, config_hash, created_at):
    rng = np.random.default_rng(cfg["random_seed"])
    src, parent = ro(source_path), ro(parent_path)
    dst = sqlite3.connect(output_path)
    for table in TABLES:
        ddl = src.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name=?", (table,)).fetchone()[0]
        dst.execute(ddl)
    dst.executescript("""
      CREATE TABLE _generation_metadata(key TEXT PRIMARY KEY, value TEXT);
      CREATE TABLE _source_session_map(synthetic_session_id TEXT PRIMARY KEY, source_session_id TEXT NOT NULL, replication_index INTEGER NOT NULL);
      CREATE TABLE _source_id_map(entity_type TEXT NOT NULL, synthetic_id TEXT PRIMARY KEY, source_id TEXT NOT NULL, synthetic_session_id TEXT NOT NULL);
    """)
    for table in ["hotel", "room"]:
        write_rows(dst, table, [dict(x) for x in src.execute(f'SELECT * FROM "{table}"')])

    searches = pd.read_sql_query("SELECT * FROM search", src)
    filters = pd.read_sql_query("SELECT * FROM search_filter", src)
    results = pd.read_sql_query("SELECT * FROM search_result", src)
    events = pd.read_sql_query("SELECT * FROM event", src)
    sessions = np.array(sorted(searches.session_id.unique()))
    q, remainder = divmod(cfg["n_sessions"], len(sessions))
    chosen = np.tile(sessions, q)
    if remainder:
        chosen = np.concatenate([chosen, rng.choice(sessions, size=remainder, replace=False)])
    chosen = rng.permutation(chosen)

    valid_durations = []
    for row in searches.itertuples():
        try:
            nights = (datetime.fromisoformat(row.checkout_date) - datetime.fromisoformat(row.checkin_date)).days
            if nights > 0:
                valid_durations.append(nights)
        except Exception:
            pass

    users, out_s, out_f, out_r, out_e, session_map, id_map = [], [], [], [], [], [], []
    replication_counter = defaultdict(int)
    for i, source_session in enumerate(chosen, 1):
        replication_counter[source_session] += 1
        uid, sid = f"SYN_U{i:04d}", f"SYN_S{i:04d}"
        base = parse_ts(cfg["base_time"]) + timedelta(days=i // 4, minutes=i % 4)
        users.append({"user_id": uid, "user_name": f"합성사용자_{i:06d}", "age_group": None,
                      "email": f"synthetic{i:06d}@example.invalid", "signup_at": format_ts(base - timedelta(days=30)),
                      "data_origin": "synthetic_augmentation"})
        sg = searches[searches.session_id == source_session].sort_values(["search_time", "search_id"], na_position="last")
        event_rows = replicated_event_rows(events, results, source_session)
        candidates = [parse_ts(x) for x in sg.search_time]
        candidates += [parse_ts(row.event_at) for _, row in event_rows]
        candidates = [x for x in candidates if x is not None]
        if not candidates:
            raise ValueError(f"no valid session-origin candidate: {source_session}")
        session_origin = min(candidates)
        qmap, fmap = {}, {}
        session_map.append((sid, source_session, replication_counter[source_session]))

        for j, row in enumerate(sg.itertuples(), 1):
            qid = f"SYN_Q{i:04d}_{j:03d}"
            qmap[row.search_id] = qid
            rec = row._asdict(); rec.pop("Index", None)
            rec.update(search_id=qid, session_id=sid, search_time=shifted(row.search_time, base, session_origin), data_origin="synthetic_augmentation")
            try:
                checkin, checkout = datetime.fromisoformat(str(row.checkin_date)), datetime.fromisoformat(str(row.checkout_date))
                if checkout <= checkin:
                    rec["checkout_date"] = (checkin + timedelta(days=int(rng.choice(valid_durations)))).date().isoformat()
            except Exception:
                pass
            out_s.append(rec); id_map.append(("search", qid, row.search_id, sid))
            fr = filters[filters.search_id == row.search_id].iloc[0].to_dict()
            source_fid = fr["search_filter_id"]
            fid = f"SYN_F{i:04d}_{j:03d}"; fmap[source_fid] = fid
            fr.update(search_filter_id=fid, search_id=qid, data_origin="synthetic_augmentation")
            out_f.append(fr); id_map.append(("search_filter", fid, source_fid, sid))
            for k, result in enumerate(results[results.search_id == row.search_id].sort_values("result_rank").to_dict("records"), 1):
                rid = f"SYN_R{i:04d}_{j:03d}_{k:04d}"
                source_rid = result["search_result_id"]
                result.update(search_result_id=rid, search_id=qid, data_origin="synthetic_augmentation")
                out_r.append(result); id_map.append(("search_result", rid, source_rid, sid))

        for ordinal, row in event_rows:
            rec = row._asdict(); rec.pop("Index", None)
            eid = f"SYN_E{i:04d}_{ordinal:05d}"
            rec.update(event_id=eid, session_id=sid, user_id=uid, search_id=qmap.get(row.search_id),
                       search_filter_id=fmap.get(row.search_filter_id), event_at=shifted(row.event_at, base, session_origin),
                       session_end_time=shifted(row.session_end_time, base, session_origin), data_origin="synthetic_augmentation")
            out_e.append(rec); id_map.append(("event", eid, row.event_id, sid))

    write_rows(dst, "user", users); write_rows(dst, "search", out_s); write_rows(dst, "search_filter", out_f)
    write_rows(dst, "search_result", out_r); write_rows(dst, "event", out_e)
    dst.executemany("INSERT INTO _source_session_map VALUES(?,?,?)", session_map)
    dst.executemany("INSERT INTO _source_id_map VALUES(?,?,?,?)", id_map)
    metadata = {
        "scenario_id": cfg["scenario_id"], "sample_set_type": cfg["sample_set_type"],
        "generation_version": cfg["generation_version"], "parent_s0_sha256": cfg["parent_s0_sha256"],
        "source_db_sha256": cfg["source_db_sha256"], "time_origin_policy": cfg["time_origin_policy"],
        "time_origin_reason": cfg["time_origin_reason"], "random_seed": cfg["random_seed"],
        "generation_code_sha256": code_hash, "generation_config_sha256": config_hash,
        "generated_at_kst": created_at, "sqlite_source_access": "URI mode=ro; PRAGMA query_only=ON",
    }
    dst.executemany("INSERT INTO _generation_metadata VALUES(?,?)", [(k, json.dumps(v, ensure_ascii=False)) for k, v in metadata.items()])
    dst.executescript("""
      CREATE INDEX idx_r1_search_session_time ON search(session_id,search_time,search_id);
      CREATE INDEX idx_r1_event_session_time ON event(session_id,event_at,event_id);
      CREATE INDEX idx_r1_event_search_hotel ON event(search_id,hotel_id,event_type);
      CREATE INDEX idx_r1_result_search_hotel ON search_result(search_id,hotel_id);
      CREATE INDEX idx_r1_map_source ON _source_id_map(entity_type,source_id);
    """)
    dst.commit(); dst.close(); src.close(); parent.close()


def table_counts(conn):
    return {t: conn.execute(f'SELECT COUNT(*) FROM "{t}"').fetchone()[0] for t in TABLES}


def duplicate_pk_failures(conn):
    failures = 0
    rows = []
    for t in TABLES:
        pk = [r[1] for r in conn.execute(f'PRAGMA table_info("{t}")') if r[5]]
        n = 0 if not pk else conn.execute(f'SELECT COUNT(*) FROM (SELECT {",".join(pk)},COUNT(*) n FROM "{t}" GROUP BY {",".join(pk)} HAVING n>1)').fetchone()[0]
        failures += n; rows.append({"table": t, "pk_columns": ",".join(pk), "duplicate_groups": n})
    return failures, rows


def fk_orphans(conn):
    total, rows = 0, []
    for child in TABLES:
        for fk in conn.execute(f'PRAGMA foreign_key_list("{child}")'):
            parent, from_col, to_col = fk[2], fk[3], fk[4]
            sql = f'SELECT COUNT(*) FROM "{child}" c LEFT JOIN "{parent}" p ON c."{from_col}"=p."{to_col}" WHERE c."{from_col}" IS NOT NULL AND p."{to_col}" IS NULL'
            n = conn.execute(sql).fetchone()[0]
            total += n; rows.append({"child": child, "column": from_col, "parent": parent, "parent_column": to_col, "orphans": n})
    return total, rows


def scalar(conn, sql):
    return conn.execute(sql).fetchone()[0]


def relative_rows(source, new):
    smap = pd.read_sql_query("SELECT * FROM _source_session_map", new)
    imap = pd.read_sql_query("SELECT * FROM _source_id_map", new)
    src_s = pd.read_sql_query("SELECT search_id,session_id,search_time FROM search", source)
    new_s = pd.read_sql_query("SELECT search_id,session_id,search_time FROM search", new)
    src_e = pd.read_sql_query("SELECT event_id,session_id,search_id,hotel_id,event_type,event_at,session_end_time FROM event", source)
    new_e = pd.read_sql_query("SELECT event_id,session_id,search_id,hotel_id,event_type,event_at,session_end_time FROM event", new)
    qmap = imap[imap.entity_type == "search"][["synthetic_id", "source_id"]]
    emap = imap[imap.entity_type == "event"][["synthetic_id", "source_id"]]
    ns = new_s.merge(qmap, left_on="search_id", right_on="synthetic_id").merge(src_s, left_on="source_id", right_on="search_id", suffixes=("_new", "_src"))
    ne = new_e.merge(emap, left_on="event_id", right_on="synthetic_id").merge(src_e, left_on="source_id", right_on="event_id", suffixes=("_new", "_src"))
    for df, cols in [(ns, ["search_time_new", "search_time_src"]), (ne, ["event_at_new", "event_at_src", "session_end_time_new", "session_end_time_src"])]:
        for col in cols: df[col] = df[col].map(parse_ts)

    checks = []
    def add(name, source_values, new_values):
        source_values, new_values = list(source_values), list(new_values)
        comparable = [(a, b) for a, b in zip(source_values, new_values) if a is not None and b is not None]
        mismatches = sum(abs(a - b) > 1e-9 for a, b in comparable)
        checks.append({"invariant": name, "comparable_n": len(comparable), "mismatch_n": mismatches, "max_abs_error_seconds": max([abs(a-b) for a,b in comparable] or [0])})

    src_gaps, new_gaps = [], []
    for _, group in ns.sort_values(["session_id_new", "search_time_new", "search_id_new"]).groupby("session_id_new"):
        src_times, new_times = group.search_time_src.tolist(), group.search_time_new.tolist()
        src_gaps += [(b-a).total_seconds() for a,b in zip(src_times,src_times[1:])]
        new_gaps += [(b-a).total_seconds() for a,b in zip(new_times,new_times[1:])]
    add("adjacent_search_gap", src_gaps, new_gaps)
    src_gaps, new_gaps = [], []
    for _, group in ne.sort_values(["session_id_new", "event_at_new", "event_id_new"]).groupby("session_id_new"):
        st, nt = group.event_at_src.tolist(), group.event_at_new.tolist()
        src_gaps += [(b-a).total_seconds() for a,b in zip(st,st[1:]) if a and b]
        new_gaps += [(b-a).total_seconds() for a,b in zip(nt,nt[1:]) if a and b]
    add("adjacent_event_gap", src_gaps, new_gaps)

    search_lookup_src = src_s.set_index("search_id").search_time.map(parse_ts).to_dict()
    search_lookup_new = new_s.set_index("search_id").search_time.map(parse_ts).to_dict()
    source_q_for_new = qmap.set_index("synthetic_id").source_id.to_dict()
    linked_src, linked_new = [], []
    for row in ne.itertuples():
        if row.search_id_new and row.event_at_src and row.event_at_new:
            sq = source_q_for_new.get(row.search_id_new)
            if sq and search_lookup_src.get(sq) and search_lookup_new.get(row.search_id_new):
                linked_src.append((row.event_at_src-search_lookup_src[sq]).total_seconds())
                linked_new.append((row.event_at_new-search_lookup_new[row.search_id_new]).total_seconds())
    add("event_minus_linked_search", linked_src, linked_new)

    next_src, next_new = [], []
    for _, group in ns.sort_values(["session_id_new", "search_time_new", "search_id_new"]).groupby("session_id_new"):
        rows = list(group.itertuples())
        next_by_q = {rows[i].search_id_new: rows[i+1] for i in range(len(rows)-1)}
        evg = ne[ne.session_id_new == rows[0].session_id_new]
        for ev in evg.itertuples():
            nxt = next_by_q.get(ev.search_id_new)
            if nxt and ev.event_at_src and ev.event_at_new:
                next_src.append((ev.event_at_src-nxt.search_time_src).total_seconds())
                next_new.append((ev.event_at_new-nxt.search_time_new).total_seconds())
    add("event_minus_next_search", next_src, next_new)

    end_ev_src, end_ev_new, end_s_src, end_s_new = [], [], [], []
    for sid, evg in ne.groupby("session_id_new"):
        ends = [(r.session_end_time_src, r.session_end_time_new) for r in evg.itertuples() if r.session_end_time_src and r.session_end_time_new]
        if not ends: continue
        end_src, end_new = ends[0]
        last_src = max(x for x in evg.event_at_src if x); last_new = max(x for x in evg.event_at_new if x)
        sg = ns[ns.session_id_new == sid]
        last_ss, last_sn = max(x for x in sg.search_time_src if x), max(x for x in sg.search_time_new if x)
        end_ev_src.append((end_src-last_src).total_seconds()); end_ev_new.append((end_new-last_new).total_seconds())
        end_s_src.append((end_src-last_ss).total_seconds()); end_s_new.append((end_new-last_sn).total_seconds())
    add("session_end_minus_last_event", end_ev_src, end_ev_new)
    add("session_end_minus_last_search", end_s_src, end_s_new)

    dc_src, dc_new = [], []
    for (_, qid, hotel), group in ne[ne.event_type_new.isin(HOTEL_EVENTS)].groupby(["session_id_new", "search_id_new", "hotel_id_new"], dropna=False):
        clicks = group[group.event_type_new == "hotel_click"]
        details = group[group.event_type_new == "hotel_detail_view"]
        if clicks.empty or details.empty: continue
        c, d = clicks.sort_values("event_at_new").iloc[0], details.sort_values("event_at_new").iloc[0]
        dc_src.append((d.event_at_src-c.event_at_src).total_seconds()); dc_new.append((d.event_at_new-c.event_at_new).total_seconds())
    add("detail_minus_click_same_search_hotel", dc_src, dc_new)
    return checks


def cross_stream(conn, label):
    searches = pd.read_sql_query("SELECT search_id,session_id,search_time FROM search", conn)
    events = pd.read_sql_query("SELECT event_id,session_id,search_id,event_type,event_at FROM event WHERE event_type NOT LIKE 'booking_%'", conn)
    searches["t"] = searches.search_time.map(parse_ts); events["t"] = events.event_at.map(parse_ts)
    next_map = {}
    for _, g in searches.sort_values(["session_id", "t", "search_id"]).groupby("session_id"):
        rs = list(g.itertuples())
        for i, r in enumerate(rs): next_map[r.search_id] = None if i+1 == len(rs) else rs[i+1].t
    categories, over, affected_sessions, affected_searches = defaultdict(int), [], set(), set()
    by_type = defaultdict(lambda: defaultdict(int))
    search_ids = set(searches.search_id)
    for e in events.itertuples():
        if not e.search_id or e.search_id not in search_ids or e.t is None:
            cat = "키 부족"
        elif next_map[e.search_id] is None:
            cat = "다음 검색 없음"
        else:
            delta = (e.t-next_map[e.search_id]).total_seconds()
            cat = "다음 검색 이전" if delta < 0 else ("다음 검색과 동일 시각" if delta == 0 else "다음 검색 이후")
            if delta >= 0:
                affected_sessions.add(e.session_id); affected_searches.add(e.search_id); over.append(delta)
        categories[cat] += 1; by_type[e.event_type][cat] += 1
    comparable = sum(categories[x] for x in ["다음 검색 이전", "다음 검색과 동일 시각", "다음 검색 이후"])
    overlap = categories["다음 검색과 동일 시각"] + categories["다음 검색 이후"]
    summary = {"dataset": label, **categories, "비교 가능": comparable, "중첩": overlap,
               "중첩률": overlap/comparable if comparable else None, "영향 세션 수": len(affected_sessions), "영향 검색 수": len(affected_searches),
               "초과시간 중앙값": float(np.median(over)) if over else None, "Q1": float(np.quantile(over,.25)) if over else None,
               "Q3": float(np.quantile(over,.75)) if over else None, "P90": float(np.quantile(over,.9)) if over else None, "최대": max(over) if over else None}
    detail = [{"dataset": label, "event_type": typ, **counts} for typ, counts in sorted(by_type.items())]
    return summary, detail


def trace_samples(source, parent, new):
    smap = pd.read_sql_query("SELECT * FROM _source_session_map", new)
    imap = pd.read_sql_query("SELECT * FROM _source_id_map", new)
    syn_to_src_q = imap[imap.entity_type == "search"].set_index("synthetic_id").source_id.to_dict()
    syn_to_src_e = imap[imap.entity_type == "event"].set_index("synthetic_id").source_id.to_dict()
    src_search = pd.read_sql_query("SELECT search_id,session_id,search_time FROM search", source)
    old_search = pd.read_sql_query("SELECT search_id,session_id,search_time FROM search", parent)
    new_search = pd.read_sql_query("SELECT search_id,session_id,search_time FROM search", new)
    src_event = pd.read_sql_query("SELECT event_id,search_id,event_type,event_at FROM event", source).set_index("event_id")
    old_event = pd.read_sql_query("SELECT event_id,session_id,search_id,event_type,event_at FROM event WHERE event_type NOT LIKE 'booking_%'", parent).set_index("event_id")
    new_event = pd.read_sql_query("SELECT event_id,session_id,search_id,event_type,event_at FROM event", new)
    src_search["t"] = src_search.search_time.map(parse_ts); old_search["t"] = old_search.search_time.map(parse_ts); new_search["t"] = new_search.search_time.map(parse_ts)
    src_next = {}; old_next = {}; new_next = {}
    for frame, dest in [(src_search,src_next),(old_search,old_next),(new_search,new_next)]:
        for _,g in frame.sort_values(["session_id","t","search_id"]).groupby("session_id"):
            rr=list(g.itertuples())
            for i,r in enumerate(rr[:-1]): dest[r.search_id]=rr[i+1].t
    candidates=[]
    for ne in new_event.itertuples():
        if ne.event_id not in syn_to_src_e or ne.search_id not in syn_to_src_q: continue
        if ne.event_id not in old_event.index: continue
        oe=old_event.loc[ne.event_id]; src_eid=syn_to_src_e[ne.event_id]; src_qid=syn_to_src_q[ne.search_id]
        if src_eid not in src_event.index or src_qid not in src_next or ne.search_id not in new_next or oe.search_id not in old_next: continue
        st=parse_ts(src_event.loc[src_eid].event_at); ot=parse_ts(oe.event_at); nt=parse_ts(ne.event_at)
        if not st or not ot or not nt: continue
        sd=(st-src_next[src_qid]).total_seconds(); od=(ot-old_next[oe.search_id]).total_seconds(); nd=(nt-new_next[ne.search_id]).total_seconds()
        candidates.append({"source_event_id":src_eid,"old_event_id":ne.event_id,"new_event_id":ne.event_id,"source_search_id":src_qid,
            "old_search_id":oe.search_id,"new_search_id":ne.search_id,"event_type":ne.event_type,
            "source_event_time":format_ts(st),"old_event_time":format_ts(ot),"new_event_time":format_ts(nt),
            "source_next_search_time":format_ts(src_next[src_qid]),"old_next_search_time":format_ts(old_next[oe.search_id]),"new_next_search_time":format_ts(new_next[ne.search_id]),
            "source_relative_seconds":sd,"old_relative_seconds":od,"new_relative_seconds":nd,
            "before_judgment":"before" if od<0 else ("equal" if od==0 else "after"),"after_judgment":"before" if nd<0 else ("equal" if nd==0 else "after")})
    def choose(label, pred, key=None):
        vals=[x for x in candidates if pred(x)]
        if not vals:return {"sample_type":label,"status":"NOT_FOUND"}
        x=(max(vals,key=key) if key else vals[0]).copy();x.update(sample_type=label,status="TRACED");return x
    return [
      choose("이전→이후 왜곡",lambda x:x["source_relative_seconds"]<0 and x["old_relative_seconds"]>=0),
      choose("최대 초과",lambda x:x["old_relative_seconds"]>=0,key=lambda x:x["old_relative_seconds"]),
      choose("hotel_click",lambda x:x["event_type"]=="hotel_click"),
      choose("hotel_detail_view",lambda x:x["event_type"]=="hotel_detail_view"),
      choose("동일 시각",lambda x:x["source_relative_seconds"]==0),
      choose("정상 보존",lambda x:(x["source_relative_seconds"] < 0) == (x["old_relative_seconds"] < 0) and x["source_relative_seconds"]==x["new_relative_seconds"]),
    ]


def qa_all(source_path, parent_path, new_path):
    source, parent, new = ro(source_path), ro(parent_path), ro(new_path)
    source_counts, parent_counts, new_counts = table_counts(source), table_counts(parent), table_counts(new)
    dup_total, dup_rows = duplicate_pk_failures(new); orphan_total, orphan_rows = fk_orphans(new)
    checks = [
      ("integrity_check", "ok", scalar(new,"PRAGMA integrity_check")),
      ("USER 1000",1000,new_counts["user"]),("session 1000",1000,scalar(new,"SELECT COUNT(DISTINCT session_id) FROM search")),
      ("SEARCH 6900",6900,new_counts["search"]),("SEARCH_FILTER 6900",6900,new_counts["search_filter"]),
      ("SEARCH_RESULT same as S0",parent_counts["search_result"],new_counts["search_result"]),("EVENT same as S0",parent_counts["event"],new_counts["event"]),
      ("BOOKING 0",0,new_counts["booking"]),("PK duplicate 0",0,dup_total),("FK orphan 0",0,orphan_total),
      ("SEARCH-FILTER 1:1 failures",0,scalar(new,"SELECT COUNT(*) FROM search s LEFT JOIN (SELECT search_id,COUNT(*) n FROM search_filter GROUP BY search_id) f USING(search_id) WHERE COALESCE(f.n,0)<>1")),
      ("result count mismatch",0,scalar(new,"SELECT COUNT(*) FROM search s LEFT JOIN (SELECT search_id,COUNT(*) n FROM search_result GROUP BY search_id) r USING(search_id) WHERE s.total_result_count<>COALESCE(r.n,0)")),
      ("unexposed click/detail",0,scalar(new,"SELECT COUNT(*) FROM event e WHERE e.event_type IN ('hotel_click','hotel_detail_view') AND NOT EXISTS(SELECT 1 FROM search_result r WHERE r.search_id=e.search_id AND r.hotel_id=e.hotel_id)")),
      ("zero-result exposure/click/detail",0,scalar(new,"SELECT COUNT(*) FROM event e JOIN search s ON e.search_id=s.search_id WHERE s.total_result_count=0 AND e.event_type IN ('hotel_impression','hotel_click','hotel_detail_view')")),
      ("negative SEARCH gap",0,scalar(new,"WITH x AS(SELECT search_time,LAG(search_time) OVER(PARTITION BY session_id ORDER BY search_time,search_id) p FROM search) SELECT COUNT(*) FROM x WHERE datetime(replace(search_time,' KST',''))<datetime(replace(p,' KST',''))")),
      ("negative EVENT gap",0,scalar(new,"WITH x AS(SELECT event_at,LAG(event_at) OVER(PARTITION BY session_id ORDER BY event_at,event_id) p FROM event) SELECT COUNT(*) FROM x WHERE datetime(replace(event_at,' KST',''))<datetime(replace(p,' KST',''))")),
      ("detail earlier than click",0,scalar(new,"SELECT COUNT(*) FROM event d WHERE d.event_type='hotel_detail_view' AND EXISTS(SELECT 1 FROM event c WHERE c.event_type='hotel_click' AND c.search_id=d.search_id AND c.hotel_id=d.hotel_id) AND datetime(replace(d.event_at,' KST','')) < (SELECT MIN(datetime(replace(c.event_at,' KST',''))) FROM event c WHERE c.event_type='hotel_click' AND c.search_id=d.search_id AND c.hotel_id=d.hotel_id)")),
      ("session_end before last EVENT",0,scalar(new,"WITH x AS(SELECT session_id,MAX(datetime(replace(event_at,' KST',''))) last_event,MAX(datetime(replace(session_end_time,' KST',''))) end_time FROM event GROUP BY session_id) SELECT COUNT(*) FROM x WHERE end_time<last_event")),
      ("session_end before last SEARCH",0,scalar(new,"WITH e AS(SELECT session_id,MAX(datetime(replace(session_end_time,' KST',''))) end_time FROM event GROUP BY session_id),s AS(SELECT session_id,MAX(datetime(replace(search_time,' KST',''))) last_search FROM search GROUP BY session_id) SELECT COUNT(*) FROM e JOIN s USING(session_id) WHERE end_time<last_search")),
      ("timestamp parse failures",0,scalar(new,"SELECT (SELECT COUNT(*) FROM search WHERE search_time IS NOT NULL AND datetime(replace(search_time,' KST','')) IS NULL)+(SELECT COUNT(*) FROM event WHERE event_at IS NOT NULL AND datetime(replace(event_at,' KST','')) IS NULL)+(SELECT COUNT(*) FROM event WHERE session_end_time IS NOT NULL AND datetime(replace(session_end_time,' KST','')) IS NULL)")),
    ]
    check_rows=[{"check":n,"expected":e,"actual":a,"status":"PASS" if a==e else "FAIL"} for n,e,a in checks]
    invariants=relative_rows(source,new)
    cross_summaries=[];cross_details=[]
    for conn,label in [(source,"ORIGINAL_296"),(parent,"S0_1000"),(new,"S0_TIME_ALIGNED_1000")]:
        s,d=cross_stream(conn,label);cross_summaries.append(s);cross_details.extend(d)
    traces=trace_samples(source,parent,new)
    source.close();parent.close();new.close()
    return {"counts":{"source":source_counts,"parent_s0":parent_counts,"new":new_counts},"qa_checks":check_rows,
            "pk_duplicates":dup_rows,"fk_orphans":orphan_rows,"time_invariants":invariants,
            "cross_stream_summary":cross_summaries,"cross_stream_by_event_type":cross_details,"deterministic_samples":traces}


def write_xlsx(path, payload):
    with pd.ExcelWriter(path,engine="openpyxl") as writer:
        pd.DataFrame([{"dataset":k,**v} for k,v in payload["counts"].items()]).to_excel(writer,"규모",index=False)
        for key,sheet in [("qa_checks","QA"),("pk_duplicates","PK중복"),("fk_orphans","FK고아"),("time_invariants","시간불변조건"),
                          ("cross_stream_summary","교차스트림요약"),("cross_stream_by_event_type","이벤트유형별"),("deterministic_samples","결정적표본추적")]:
            pd.DataFrame(payload[key]).to_excel(writer,sheet,index=False)
    wb=load_workbook(path,read_only=True); expected=8; actual=len(wb.sheetnames);wb.close();assert actual==expected


def main():
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    ap=argparse.ArgumentParser();ap.add_argument("--config",type=Path,required=True);ap.add_argument("--source-db",type=Path,required=True);ap.add_argument("--parent-s0",type=Path,required=True);ap.add_argument("--output-dir",type=Path,required=True)
    a=ap.parse_args(); cfg=json.loads(a.config.read_text(encoding="utf-8"));out=a.output_dir.resolve()
    if not out.is_dir(): raise FileNotFoundError("output directory must already exist")
    source_before,parent_before=sha256(a.source_db),sha256(a.parent_s0)
    assert source_before==cfg["source_db_sha256"] and parent_before==cfg["parent_s0_sha256"]
    stamp="260905_1624_01"; prefix="호텔검색_관측형합성1000명_"
    files={
      "db":out/f"{prefix}시간축보정데이터_{stamp}.sqlite","xlsx":out/f"{prefix}생성직후QA결과_{stamp}.xlsx",
      "json":out/f"{prefix}시간축보정결과_{stamp}.json","trace":out/f"{prefix}결정적표본추적표_{stamp}.xlsx",
      "runlog":out/f"{prefix}시간축보정실행로그_{stamp}.md","decision":out/f"{prefix}변경판단로그_{stamp}.md",
      "manifest":out/f"{prefix}SHA256매니페스트_{stamp}.json"}
    for p in files.values():
        if p.exists(): raise FileExistsError(p)
    created=datetime.now(KST).isoformat();code_hash=sha256(Path(__file__));config_hash=sha256(a.config)
    build(a.source_db,a.parent_s0,files["db"],cfg,code_hash,config_hash,created)
    payload=qa_all(a.source_db,a.parent_s0,files["db"])
    payload.update({"step":"STEP_R1","created_at_kst":created,"config":cfg,"input_hashes_before":{"source":source_before,"parent_s0":parent_before}})
    payload["input_hashes_after"]={"source":sha256(a.source_db),"parent_s0":sha256(a.parent_s0)}
    payload["inputs_immutable"]=payload["input_hashes_before"]==payload["input_hashes_after"]
    failed=[x for x in payload["qa_checks"] if x["status"]!="PASS"]+[x for x in payload["time_invariants"] if x["mismatch_n"]]
    payload["status"]="PASS" if not failed and payload["inputs_immutable"] else "FAIL"
    write_xlsx(files["xlsx"],payload)
    with pd.ExcelWriter(files["trace"],engine="openpyxl") as w: pd.DataFrame(payload["deterministic_samples"]).to_excel(w,"결정적표본",index=False)
    files["json"].write_text(json.dumps(payload,ensure_ascii=False,indent=2,default=str),encoding="utf-8")
    cs=payload["cross_stream_summary"]
    files["runlog"].write_text("\n".join(["# STEP R1 실행 로그",f"- 생성시각(KST): `{created}`",f"- 원본 접근: URI mode=ro, PRAGMA query_only=ON",f"- 원점: `{cfg['time_origin_policy']}`",f"- 공식: `new_time = base_i + (old_time - session_origin)`",f"- seed: `{cfg['random_seed']}`",f"- 결과: **STEP R1={payload['status']}**"]),encoding="utf-8")
    files["decision"].write_text("\n".join(["# STEP R1 변경 판단 로그","- 수정 전 원인: SEARCH와 EVENT에 서로 다른 원점을 적용하여 교차 스트림 상대시간이 소실됨.","- 수정 범위: SEARCH.search_time, EVENT.event_at, EVENT.session_end_time의 시간 이동 원점만 공통 session_origin으로 변경.","- 선택 원점: 유효 SEARCH.search_time과 실제 복제 대상 EVENT.event_at의 최솟값.","- 이유: 검색 전 session_start를 유지하고 모든 내부·교차 시간차를 동일 평행이동으로 정확히 보존.","- Jittering/신규 시간분포: 적용하지 않음.",f"- 원본/S0/신규 중첩: {cs[0]['중첩']}/{cs[0]['비교 가능']}, {cs[1]['중첩']}/{cs[1]['비교 가능']}, {cs[2]['중첩']}/{cs[2]['비교 가능']}",f"- 최종 판정: **STEP R1={payload['status']}**"]),encoding="utf-8")
    artifacts=[Path(__file__),a.config,*files.values()]
    manifest={"created_at_kst":datetime.now(KST).isoformat(),"inputs":{"source_db":{"path":str(a.source_db.resolve()),"sha256":sha256(a.source_db)},"parent_s0":{"path":str(a.parent_s0.resolve()),"sha256":sha256(a.parent_s0)}},"outputs":[]}
    for p in artifacts:
        if p.exists() and p!=files["manifest"]: manifest["outputs"].append({"path":str(p.resolve()),"size_bytes":p.stat().st_size,"sha256":sha256(p)})
    files["manifest"].write_text(json.dumps(manifest,ensure_ascii=False,indent=2),encoding="utf-8")
    assert sha256(a.source_db)==source_before and sha256(a.parent_s0)==parent_before
    print(json.dumps({"status":payload["status"],"output_db":str(files["db"]),"counts":payload["counts"],"cross_stream":cs,"failed":failed},ensure_ascii=False,indent=2))
    if payload["status"]!="PASS":sys.exit(1)


if __name__=="__main__": main()
