#!/usr/bin/env python3
"""STEP2 full hypothesis calculation for approved ORIGINAL_296 and S0_1000.

The approved execution context is the sole entry input. Both SQLite databases
are opened with URI mode=ro and PRAGMA query_only=ON. This program never runs
the generator and never creates synthetic, stress, oversampled, A3, or booking
data. BOOKING is used only for the entry row-count gate.
"""
from __future__ import annotations

import argparse
import hashlib
import importlib.metadata
import json
import math
import platform
import sqlite3
import sys
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from scipy.stats import fisher_exact, mannwhitneyu

KST = ZoneInfo("Asia/Seoul")
BUNDLE_TS = "260904_1149"
EXPECTED_CONTEXT_SHA256 = "b1558b92f5dba089316dd9cae338f7b9a5a72685b878d81c9ffb15fe7f0723e9"
EXPECTED_CORE = {
    "ORIGINAL_296": {
        "ZERO_RESULT_RATE": (147, 296), "ZERO_FOLLOWUP_RATE": (140, 147),
        "B3_IMMEDIATE_RECOVERY": (24, 140), "B3_SESSION_FINAL_RECOVERY": (21, 28),
        "SEARCH_HOTEL_CLICK_RATE": (45, 296), "DIAG_FIRST_SEARCH_ZERO_RECOVERY": (4, 6),
    },
    "S0_1000": {
        "ZERO_RESULT_RATE": (3434, 6900), "ZERO_FOLLOWUP_RATE": (3271, 3434),
        "B3_IMMEDIATE_RECOVERY": (558, 3271), "B3_SESSION_FINAL_RECOVERY": (488, 651),
        "SEARCH_HOTEL_CLICK_RATE": (1052, 6900), "DIAG_FIRST_SEARCH_ZERO_RECOVERY": (92, 139),
    },
}
REGIONS = ["Tokyo", "Osaka", "Kyoto", "Sapporo", "Fukuoka", "UNKNOWN"]
INTENTS = ["LOCATION_ONLY", "PRICE", "QUALITY_FILTER", "AMENITY", "MIXED"]
H3_ORDER = ["동일조건 반복", "조건 완화", "검색어 수정", "지역 변경", "조건 강화", "혼합 변경"]
SEGMENT_ORDER = ["직접 성공", "결과 노출·미선택", "재검색 회복", "지속 실패"]


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for block in iter(lambda: f.read(1024 * 1024), b""):
            h.update(block)
    return h.hexdigest()


def lf_sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes().replace(b"\r\n", b"\n")).hexdigest()


def open_ro(path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(path.resolve().as_uri() + "?mode=ro", uri=True)
    conn.execute("PRAGMA query_only=ON")
    assert conn.execute("PRAGMA query_only").fetchone()[0] == 1
    assert conn.execute("PRAGMA integrity_check").fetchone()[0] == "ok"
    return conn


def resolve(root: Path, value: str) -> Path:
    p = Path(value)
    if p.is_absolute() and p.exists():
        return p.resolve()
    return (root / value.replace("/", "\\")).resolve(strict=True)


def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--context", type=Path, required=True)
    p.add_argument("--output-dir", type=Path, required=True)
    p.add_argument("--version", default="01")
    return p.parse_args()


def record(assertions, check, actual, expected, passed=None, detail=""):
    ok = actual == expected if passed is None else bool(passed)
    assertions.append({"check": check, "actual": str(actual), "expected": str(expected),
                       "status": "PASS" if ok else "FAIL", "detail": detail})
    if not ok:
        raise AssertionError(f"{check}: actual={actual!r}, expected={expected!r}; {detail}")


def preflight(root: Path, context_path: Path, assertions: list[dict]):
    record(assertions, "context raw SHA-256", sha256(context_path), EXPECTED_CONTEXT_SHA256)
    ctx = json.loads(context_path.read_text(encoding="utf-8"))
    for key, expected in {
        "step1_status": "PASS", "hash_gate_status": "PASS", "BUNDLE_RUN_TS": BUNDLE_TS,
        "parent_expansion_allowed": True, "existing_s0_1000_hypothesis_analysis_allowed": True,
        "generator_execution_allowed": False, "this_workflow_10000_generation_allowed": False,
    }.items():
        record(assertions, f"context::{key}", ctx.get(key), expected)

    add_path = resolve(root, ctx["transport_normalization_addendum"]["path"])
    parent_path = resolve(root, ctx["parent_corrected_manifest"]["path"])
    record(assertions, "addendum SHA-256", sha256(add_path), ctx["transport_normalization_addendum"]["sha256"])
    record(assertions, "parent manifest SHA-256", sha256(parent_path), ctx["parent_corrected_manifest"]["sha256"])
    add = json.loads(add_path.read_text(encoding="utf-8"))
    parent = json.loads(parent_path.read_text(encoding="utf-8"))
    record(assertions, "addendum HASH_GATE", add.get("HASH_GATE"), "PASS")
    for key in ("generation_code", "generation_config"):
        record(assertions, f"{key} equivalence", add["text_transport_audit"][key]["equivalence_class"], "CRLF_TO_LF_ONLY")
    record(assertions, "parent final QA", parent.get("final_qa_status"), "PASS")
    record(assertions, "parent expansion_allowed", parent.get("expansion_allowed"), True)
    for g in ["G1", "G2", "G3", "G4", "G5"]:
        record(assertions, f"parent {g}", parent["G1_G5"][g], "PASS")
    config_info = ctx["config_hashes"]
    record(assertions, "parent seed", config_info.get("random_seed"), 20260903)

    files = {}
    for item in ctx["all_followup_inputs"]:
        path = resolve(root, item["path"]); files[item["role"]] = path
        record(assertions, f"raw hash::{item['role']}", sha256(path), item["current_raw_sha256"])
        record(assertions, f"size::{item['role']}", path.stat().st_size, item["size_bytes"])
    for role, info_key in [("code", "generation_code_hashes"), ("config", "config_hashes")]:
        info = ctx[info_key]
        record(assertions, f"LF hash::{role}", lf_sha256(files[role]), info["lf_normalized_sha256"])
        record(assertions, f"LF parent hash::{role}", lf_sha256(files[role]), info["parent_approved_sha256"])

    expected_counts = {
        "source": {"user": 89, "search": 296, "search_filter": 296, "search_result": 8555, "event": 10432, "booking": 36},
        "synthetic": {"user": 1000, "search": 6900, "booking": 0},
    }
    for role in ("source", "synthetic"):
        with open_ro(files[role]) as conn:
            for table, expected in expected_counts[role].items():
                actual = conn.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0]
                record(assertions, f"rows::{role}::{table}", actual, expected)
            sessions = conn.execute('SELECT COUNT(DISTINCT session_id) FROM search').fetchone()[0]
            record(assertions, f"sessions::{role}", sessions, 43 if role == "source" else 1000)
    return ctx, files, add_path, parent_path


def fisher_ci(a, b, c, d):
    # OR=(a/b)/(c/d), exposed group first. Fisher uses raw cells. Wald log-OR
    # CI uses raw cells unless any cell is zero, when 0.5 is added to CI cells only.
    result = fisher_exact([[a, b], [c, d]], alternative="two-sided")
    corrected = any(x == 0 for x in (a, b, c, d))
    aa, bb, cc, dd = ([x + 0.5 for x in (a, b, c, d)] if corrected else (a, b, c, d))
    log_or = math.log((aa * dd) / (bb * cc))
    se = math.sqrt(1 / aa + 1 / bb + 1 / cc + 1 / dd)
    return float(result.statistic), float(math.exp(log_or - 1.96 * se)), float(math.exp(log_or + 1.96 * se)), float(result.pvalue), corrected


def make_row(dataset_type, metric_id, analysis_unit, **kwargs):
    base = {"dataset_type": dataset_type, "metric_id": metric_id, "analysis_unit": analysis_unit,
            "group": None, "numerator": None, "denominator": None, "rate": None, "n": None,
            "test": None, "statistic": None, "p_value": None, "effect_size": None,
            "ci_95_low": None, "ci_95_high": None, "interpretation": None, "limitation": None}
    base.update(kwargs); return base


def city(value, normalize_text):
    text = normalize_text(value)
    if text is None: return "UNKNOWN"
    for raw, label in [("tokyo", "Tokyo"), ("osaka", "Osaka"), ("kyoto", "Kyoto"), ("sapporo", "Sapporo"), ("fukuoka", "Fukuoka")]:
        if text == raw or text.startswith(raw + " ·"): return label
    return "UNKNOWN"


def calculate(db: Path, dataset_type: str, marts):
    with open_ro(db) as conn:
        search, filt, event = marts.load_sources(conn)
        search, filt, event = marts.coerce_types(search, filt, event)
        base = marts.build_search_base(search, filt)
        flags = marts.build_event_search_flags(base, event)
        ordered = marts.build_ordered_searches(base)
        transitions = marts.build_zero_transitions(ordered)
        summary = marts.build_session_summary(ordered, flags)
        assignments = marts.build_session_segments(summary)

    # Any-zero session is recovered if at least one positive search occurs after
    # any zero search, not merely when the first search is zero or final row > 0.
    recovered = {}
    for sid, g in ordered.groupby("session_id", sort=False):
        vals = g["total_result_count"].astype(int).tolist()
        recovered[sid] = any(v == 0 and any(x > 0 for x in vals[i + 1:]) for i, v in enumerate(vals))
    summary["any_zero_later_positive"] = summary["session_id"].map(recovered).astype(bool)

    flag_map = flags.set_index("search_id")["has_click"]
    total, zeros = len(base), int(base.total_result_count.eq(0).sum())
    has_next = ordered.next_search_id.notna()
    zero_mask = ordered.total_result_count.eq(0)
    any_zero = summary.experienced_zero
    first_zero = summary.first_search_zero
    core_pairs = {
        "ZERO_RESULT_RATE": (zeros, total),
        "ZERO_FOLLOWUP_RATE": (int((zero_mask & has_next).sum()), zeros),
        "B3_IMMEDIATE_RECOVERY": (int(transitions.next_recovered.sum()), len(transitions)),
        "B3_SESSION_FINAL_RECOVERY": (int(summary.loc[any_zero, "any_zero_later_positive"].sum()), int(any_zero.sum())),
        "SEARCH_HOTEL_CLICK_RATE": (int(flags.has_click.sum()), total),
        "DIAG_FIRST_SEARCH_ZERO_RECOVERY": (int(summary.loc[first_zero, "subsequent_recovery"].sum()), int(first_zero.sum())),
    }
    core = [make_row(dataset_type, mid, {"ZERO_RESULT_RATE":"search", "ZERO_FOLLOWUP_RATE":"zero-result search",
        "B3_IMMEDIATE_RECOVERY":"zero-transition", "B3_SESSION_FINAL_RECOVERY":"session experiencing any zero result",
        "SEARCH_HOTEL_CLICK_RATE":"search", "DIAG_FIRST_SEARCH_ZERO_RECOVERY":"first-search-zero session"}[mid],
        numerator=num, denominator=den, rate=num/den, n=den,
        interpretation="합성은 모형 내부 진단" if dataset_type=="S0_1000" else "원본 관측자료 탐색적 결과",
        limitation="hotel_click만 상세진입 KPI; 합성 p값은 실제 모집단 증거가 아님") for mid,(num,den) in core_pairs.items()]

    a1 = []
    comparisons = [
        ("A1_AMENITY_GE3", base.amenity_count.ge(3), "amenity_count>=3", "amenity_count<3"),
        ("A1_RATING_SET", base.user_rating_min.notna(), "set", "unset"),
        ("A1_PRICE_SET", base.price.notna(), "set", "unset"),
    ]
    zero = base.total_result_count.eq(0)
    for mid, exposed, glabel, rlabel in comparisons:
        a=int((exposed&zero).sum()); b=int((exposed&~zero).sum()); c=int((~exposed&zero).sum()); d=int((~exposed&~zero).sum())
        odds, lo, hi, p, corrected = fisher_ci(a,b,c,d)
        common={"test":"Fisher exact two-sided", "statistic":odds, "p_value":p, "effect_size":odds,
                "ci_95_low":lo, "ci_95_high":hi, "interpretation":"A1 부분 채택",
                "limitation":"입력 상태를 구분할 컬럼이 없어 조건 제한 효과와 입력 품질 효과를 분리하지 못했다. 검색 반복의 독립성 한계.",
                "or_direction":f"{glabel} zero-result odds / {rlabel} zero-result odds", "ci_zero_cell_correction":corrected,
                "contingency_cells":f"a={a},b={b},c={c},d={d}"}
        a1 += [make_row(dataset_type,mid,"search",group=glabel,n=a+b,numerator=a,denominator=a+b,rate=a/(a+b),difference_pp=(a/(a+b)-c/(c+d))*100,**common),
               make_row(dataset_type,mid,"search",group=rlabel,n=c+d,numerator=c,denominator=c+d,rate=c/(c+d),difference_pp=(c/(c+d)-a/(a+b))*100,**common)]

    region_series = base.destination.map(lambda x: city(x, marts.normalize_text))
    price=base.price.notna(); rating=base.user_rating_min.notna(); amenity=base.amenity_count.fillna(0).gt(0)
    count=price.astype(int)+rating.astype(int)+amenity.astype(int)
    intent=pd.Series("LOCATION_ONLY",index=base.index); intent[count.ge(2)]="MIXED"
    intent[count.eq(1)&price]="PRICE"; intent[count.eq(1)&rating]="QUALITY_FILTER"; intent[count.eq(1)&amenity]="AMENITY"
    a2=[]
    for reg in REGIONS:
        for inte in INTENTS:
            m=region_series.eq(reg)&intent.eq(inte); n=int(m.sum()); z=int((m&zero).sum()); clicks=int(base.loc[m,"search_id"].map(flag_map).fillna(False).sum())
            a2.append(make_row(dataset_type,"A2_REGION_INTENT_ZERO_RESULT","search",group=f"{reg}|{inte}",region=reg,intent=inte,n=n,numerator=z,denominator=n,rate=z/n if n else None,
                nonzero_count=n-z,hotel_click_searches=clicks,hotel_click_rate=clicks/n if n else None,is_empty_cell=n==0,sparse_lt5=n<5,sparse_lt10=n<10,
                interpretation="기술통계",limitation="희소 셀의 우열·인과 해석 금지; 0%는 표본에서 미관측"))

    nxt=has_next; a=int((zero_mask&nxt).sum()); b=int((zero_mask&~nxt).sum()); c=int((~zero_mask&nxt).sum()); d=int((~zero_mask&~nxt).sum())
    b1_cells=(a,b,c,d)
    odds,lo,hi,p,corr=fisher_ci(a,b,c,d)
    b1=[]
    for group,num,den in [("zero_result",a,a+b),("positive_result",c,c+d)]:
        b1.append(make_row(dataset_type,"B1_IMMEDIATE_FOLLOWUP","search",group=group,n=den,numerator=num,denominator=den,rate=num/den,
            difference_pp=(a/(a+b)-c/(c+d))*100 if group=="zero_result" else (c/(c+d)-a/(a+b))*100,
            test="Fisher exact two-sided",statistic=odds,p_value=p,effect_size=odds,ci_95_low=lo,ci_95_high=hi,
            or_direction="zero-result followup odds / positive-result followup odds",contingency_cells=f"a={a},b={b},c={c},d={d}",ci_zero_cell_correction=corr,
            interpretation="B1 관련성",limitation="인과효과가 아니며 동일 세션 반복 검색의 독립성 한계"))

    yes=summary.loc[summary.experienced_zero,"search_count"].astype(float); no=summary.loc[~summary.experienced_zero,"search_count"].astype(float)
    mw=mannwhitneyu(yes,no,alternative="two-sided",method="auto")
    # Rank-biserial is positive when the first group tends to have larger values.
    rbc=2*float(mw.statistic)/(len(yes)*len(no))-1
    b2=[]
    for group,v in [("experienced_zero",yes),("no_zero_experience",no)]:
        b2.append(make_row(dataset_type,"B2_SESSION_SEARCH_COUNT","session",group=group,n=len(v),mean=float(v.mean()),median=float(v.median()),
            q1=float(v.quantile(.25)),q3=float(v.quantile(.75)),iqr=float(v.quantile(.75)-v.quantile(.25)),minimum=float(v.min()),maximum=float(v.max()),
            test="Mann–Whitney U two-sided; scipy method=auto",statistic=float(mw.statistic),p_value=float(mw.pvalue),effect_size=rbc,
            effect_direction="positive = experienced_zero has larger search count",quantile_method="pandas linear interpolation",
            interpretation="B2 비인과 진단",limitation="탐색지속성 교란 가능성; 합성 p값은 모형 내부 진단"))

    b3=[dict(x) for x in core if x["metric_id"] in {"B3_IMMEDIATE_RECOVERY","B3_SESSION_FINAL_RECOVERY","DIAG_FIRST_SEARCH_ZERO_RECOVERY"}]
    segments=[]; counts=assignments.result_segment.value_counts()
    for label in SEGMENT_ORDER:
        n=int(counts.get(label,0)); segments.append(make_row(dataset_type,"SESSION_RESULT_SEGMENT","session",group=label,n=n,numerator=n,denominator=len(summary),rate=n/len(summary),interpretation="상호배타 4개 결과 세그먼트"))

    trans=transitions.copy(); trans["next_has_click"]=trans.next_search_id.map(flag_map).fillna(False).astype(bool)
    h3=[]
    for label in H3_ORDER:
        g=trans[trans.transition_type.eq(label)]; n=len(g); success=int(g.next_recovered.sum()); click=int(g.next_has_click.sum())
        h3.append(make_row(dataset_type,"H3_TRANSITION_TYPE","zero-transition",group=label,n=n,numerator=success,denominator=n,rate=success/n if n else None,
            next_positive_count=success,next_positive_rate=success/n if n else None,next_hotel_click_count=click,next_hotel_click_rate=click/n if n else None,
            composition_rate=n/len(trans),interpretation="H3 탐색적 관계",limitation="제품 개선 인과효과로 단정 금지; hotel_click만 사용"))
    raw=[]
    for mid, table, direction in [(x[0],None,None) for x in []]: pass
    # Normalized raw 2x2 cells for exact reproduction.
    for mid, exposed, glabel, rlabel in comparisons:
        a=int((exposed&zero).sum()); b=int((exposed&~zero).sum()); c=int((~exposed&zero).sum()); d=int((~exposed&~zero).sum())
        for group,outcome,value in [(glabel,"zero_result",a),(glabel,"nonzero_result",b),(rlabel,"zero_result",c),(rlabel,"nonzero_result",d)]:
            raw.append({"dataset_type":dataset_type,"metric_id":mid,"group":group,"outcome":outcome,"count":value})
    a,b,c,d=b1_cells
    for group,outcome,value in [("zero_result","has_followup",a),("zero_result","no_followup",b),("positive_result","has_followup",c),("positive_result","no_followup",d)]:
        raw.append({"dataset_type":dataset_type,"metric_id":"B1_IMMEDIATE_FOLLOWUP","group":group,"outcome":outcome,"count":value})
    return {"core":core,"core_pairs":core_pairs,"a1":a1,"a2":a2,"b1":b1,"b2":b2,"b3":b3,"segments":segments,"h3":h3,"raw":raw,
            "counts":{"search":len(base),"sessions":len(summary),"zero_transitions":len(transitions)}}


def metric_dictionary():
    rows = [
      ("ZERO_RESULT_RATE","search","total_result_count=0","zero searches","all searches","none","Descriptive","0% means not observed"),
      ("ZERO_FOLLOWUP_RATE","zero-result search","total_result_count=0","zero searches with immediate next search","all zero searches","none","Descriptive","same-session stable order"),
      ("SEARCH_HOTEL_CLICK_RATE","search","all searches","search_id with >=1 hotel_click","all searches","none","Descriptive","hotel_detail_view excluded"),
      ("A1_AMENITY_GE3","search","amenity_count>=3 vs <3","zero-result searches","group searches","Fisher exact two-sided","OR exposed/reference","partial adoption; input-state limitation"),
      ("A1_RATING_SET","search","user_rating_min set vs unset","zero-result searches","group searches","Fisher exact two-sided","OR set/unset","partial adoption; input-state limitation"),
      ("A1_PRICE_SET","search","price set vs unset","zero-result searches","group searches","Fisher exact two-sided","OR set/unset","partial adoption; input-state limitation"),
      ("A2_REGION_INTENT_ZERO_RESULT","search","6 regions x 5 intents","zero-result searches","cell searches","descriptive only","none","sparse cells; no causal ranking"),
      ("B1_IMMEDIATE_FOLLOWUP","search","zero vs positive result","searches with immediate next","group searches","Fisher exact two-sided","OR zero/positive","association only"),
      ("B2_SESSION_SEARCH_COUNT","session","any zero experience vs none","not applicable","sessions per group","Mann-Whitney U two-sided","rank-biserial positive=zero group larger","exploration-persistence confounding"),
      ("B3_IMMEDIATE_RECOVERY","zero-transition","zero search with next","next search nonzero","zero transitions","none","none","not session-final metric"),
      ("B3_SESSION_FINAL_RECOVERY","session experiencing any zero","any zero experience","session has a later positive after any zero","any-zero sessions","none","none","separate denominator"),
      ("DIAG_FIRST_SEARCH_ZERO_RECOVERY","first-search-zero session","first search zero","later positive","first-search-zero sessions","none","none","diagnostic, not B3 core"),
      ("SESSION_RESULT_SEGMENT","session","all sessions","segment sessions","all sessions","none","none","four mutually exclusive categories"),
      ("H3_TRANSITION_TYPE","zero-transition","zero to immediate next","next positive / next hotel_click","transition type n","descriptive only","none","exploratory; no causal claim"),
    ]
    return pd.DataFrame(rows,columns=["metric_id","analysis_unit","target_condition","numerator_definition","denominator_definition","statistical_test","effect_size_direction","limitation"])


def main():
    args=parse_args(); context=args.context.resolve(strict=True); output=args.output_dir.resolve(strict=True)
    script=Path(__file__).resolve(); version=args.version.zfill(2)
    xlsx=output/f"호텔검색_관측형합성1000명_전체가설계산결과_{BUNDLE_TS}_{version}.xlsx"
    log=output/f"호텔검색_관측형합성1000명_전체가설계산실행기록_{BUNDLE_TS}_{version}.md"
    if xlsx.exists() or log.exists(): raise FileExistsError("Output exists; choose one common unused version")
    root=context.parents[3]; assertions=[]; started=datetime.now(KST).isoformat(timespec="seconds")
    ctx,files,add_path,parent_path=preflight(root,context,assertions)
    tracked=[context,add_path,parent_path,*files.values()]
    before={str(p):{"sha256":sha256(p),"mtime_ns":p.stat().st_mtime_ns,"size":p.stat().st_size} for p in tracked}
    scripts=root/'09_단계별 분석'/'1단계_원본_분석_및_가설_검증'/'scripts'
    sys.path.insert(0,str(scripts)); import build_original_296_marts as marts
    results={name:calculate(files[role],name,marts) for name,role in [("ORIGINAL_296","source"),("S0_1000","synthetic")]}
    for ds in results:
        for mid,pair in results[ds]["core_pairs"].items(): record(assertions,f"core::{ds}::{mid}",pair,EXPECTED_CORE[ds][mid])
        record(assertions,f"A2 grid rows::{ds}",len(results[ds]["a2"]),30)
        record(assertions,f"A2 n sum::{ds}",sum(x["n"] for x in results[ds]["a2"]),results[ds]["counts"]["search"])
        record(assertions,f"B3 distinct metric ids::{ds}",len({x["metric_id"] for x in results[ds]["b3"]}),3)
        record(assertions,f"segment sum::{ds}",sum(x["n"] for x in results[ds]["segments"]),results[ds]["counts"]["sessions"])
        record(assertions,f"segment labels::{ds}",len(results[ds]["segments"]),4)
        record(assertions,f"H3 n sum::{ds}",sum(x["n"] for x in results[ds]["h3"]),results[ds]["counts"]["zero_transitions"])
    record(assertions,"original segment exact",[x["n"] for x in results["ORIGINAL_296"]["segments"]],[27,10,4,2])
    record(assertions,"original H3 exact",[x["n"] for x in results["ORIGINAL_296"]["h3"]],[53,41,10,24,10,2])
    record(assertions,"synthetic H3 total",sum(x["n"] for x in results["S0_1000"]["h3"]),3271)
    for ds in results:
        for key in ("core","a1","a2","b1","b3","segments","h3"):
            for row in results[ds][key]:
                for field in ("rate","p_value"):
                    val=row.get(field)
                    if val is not None and not pd.isna(val): record(assertions,f"range::{ds}::{key}::{field}",True,True,0<=float(val)<=1)
                for field in ("statistic","effect_size","ci_95_low","ci_95_high"):
                    val=row.get(field)
                    if val is not None and isinstance(val,(int,float,np.number)): record(assertions,f"finite::{ds}::{key}::{field}",math.isfinite(float(val)),True,detail="zero-cell CI correction recorded where applicable")
    tables={
      "run_info":pd.DataFrame([{"step":"STEP2_FULL_HYPOTHESIS_CALCULATION","bundle_run_ts":BUNDLE_TS,"version":version,"started_at_kst":started,"context_path":str(context),"context_sha256":sha256(context),"source_db_sha256":sha256(files['source']),"synthetic_db_sha256":sha256(files['synthetic']),"sorting":"session_id, search_time, search_id (stable mergesort)","randomness_used":False,"generator_executed":False,"booking_kpi_used":False,"a3_excluded":True}]),
      "metric_dictionary":metric_dictionary(),
      "G3_core_metrics":pd.DataFrame(sum([results[d]["core"] for d in results],[])),
      "A1_filters":pd.DataFrame(sum([results[d]["a1"] for d in results],[])),
      "A2_region_intent":pd.DataFrame(sum([results[d]["a2"] for d in results],[])),
      "B1_followup":pd.DataFrame(sum([results[d]["b1"] for d in results],[])),
      "B2_search_count":pd.DataFrame(sum([results[d]["b2"] for d in results],[])),
      "B3_recovery":pd.DataFrame(sum([results[d]["b3"] for d in results],[])),
      "session_segments":pd.DataFrame(sum([results[d]["segments"] for d in results],[])),
      "H3_transitions":pd.DataFrame(sum([results[d]["h3"] for d in results],[])),
      "raw_contingency_tables":pd.DataFrame(sum([results[d]["raw"] for d in results],[])),
      "assertions":pd.DataFrame(assertions),
      "known_limitations":pd.DataFrame({"limitation":["원본은 296검색·43세션의 탐색적 관측 분석이다.","합성 검정은 생성 구조·방향 보존 및 코드 작동을 확인하는 모형 내부 진단이다.","합성 표본의 작은 p값은 실제 근거 강화가 아니다.","동일 세션 반복 검색으로 검색 단위 Fisher 검정에 독립성 한계가 있다.","A1 부분 채택; 입력 상태 컬럼 부재로 조건 제한 효과와 입력 품질 효과를 분리할 수 없다.","A2는 기술통계, B1은 관련성, B2는 탐색지속성 교란이 있는 비인과 진단, H3는 탐색적 관계다.","0%는 불가능이 아니라 표본에서 관측되지 않음이다.","hotel_click만 상세진입 KPI이며 hotel_detail_view는 중복 KPI로 사용하지 않는다.","A3는 제외하며 BOOKING은 KPI·가설검정에 사용하지 않는다.","합성 DB BOOKING 0행이므로 예약전환 결론을 만들 수 없다."]}),
    }
    with pd.ExcelWriter(xlsx,engine="openpyxl") as writer:
        for sheet,df in tables.items(): df.to_excel(writer,sheet_name=sheet,index=False)
    wb=load_workbook(xlsx)
    for ws in wb.worksheets:
        ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
        for cell in ws[1]: cell.font=Font(bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor="1F4E78")
        headers={c.value:i+1 for i,c in enumerate(ws[1])}
        for h in ("rate","hotel_click_rate","next_positive_rate","next_hotel_click_rate","composition_rate"):
            if h in headers:
                for col in ws.iter_cols(min_col=headers[h],max_col=headers[h],min_row=2):
                    for c in col: c.number_format="0.00%"
        for col in ws.columns:
            letter=col[0].column_letter; ws.column_dimensions[letter].width=min(50,max(10,max(len(str(c.value or "")) for c in col)+2))
    wb.save(xlsx)
    # Reopen workbook and verify required sheets, row counts, and exact core cells.
    required=list(tables); check=load_workbook(xlsx,data_only=True,read_only=True)
    record(assertions,"Excel required sheets",check.sheetnames,required)
    for name,df in tables.items(): record(assertions,f"Excel rows::{name}",check[name].max_row-1,len(df))
    core_ws=check["G3_core_metrics"]; hdr=[c.value for c in next(core_ws.iter_rows(min_row=1,max_row=1))]
    rows=[dict(zip(hdr,[c.value for c in row])) for row in core_ws.iter_rows(min_row=2)]
    for ds,expected in EXPECTED_CORE.items():
        for mid,pair in expected.items():
            rr=next(r for r in rows if r["dataset_type"]==ds and r["metric_id"]==mid)
            record(assertions,f"Excel core::{ds}::{mid}",(rr["numerator"],rr["denominator"]),pair)
    check.close()
    after={str(p):{"sha256":sha256(p),"mtime_ns":p.stat().st_mtime_ns,"size":p.stat().st_size} for p in tracked}
    record(assertions,"all inputs immutable",after,before)
    # Rewrite assertions sheet after post-save and immutable checks, then verify it.
    wb=load_workbook(xlsx); del wb["assertions"]; ws=wb.create_sheet("assertions",11)
    adf=pd.DataFrame(assertions); ws.append(list(adf.columns)); [ws.append(list(row)) for row in adf.itertuples(index=False,name=None)]
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
    for c in ws[1]: c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="1F4E78")
    wb.save(xlsx)
    finished=datetime.now(KST).isoformat(timespec="seconds")
    packages={x:importlib.metadata.version(x) for x in ["pandas","numpy","scipy","openpyxl"]}
    code_hash=sha256(script); excel_hash=sha256(xlsx)
    core_lines=[]
    for ds in ("ORIGINAL_296","S0_1000"):
        for mid,(num,den) in results[ds]["core_pairs"].items(): core_lines.append(f"- {ds} `{mid}`: {num}/{den} = {num/den:.4%}")
    log_text="# STEP2 전체 가설 계산 실행기록\n\n- 최종 판정: **STEP2=PASS**\n- 실행 시작: `"+started+"`\n- 실행 종료: `"+finished+"`\n- 실행 컨텍스트: `"+str(context)+"`\n- 컨텍스트 SHA-256: `"+sha256(context)+"`\n- SQLite: URI `mode=ro`, `PRAGMA query_only=ON`, `integrity_check=ok`\n- 정렬: `session_id, search_time, search_id` 안정 정렬\n- 난수 사용: 없음\n- 생성기 실행: 없음\n- BOOKING KPI/검정: 사용 안 함\n- A3: 제외\n\n## 실행 환경\n\n- Python: `"+platform.python_version()+"`\n- 플랫폼: `"+platform.platform()+"`\n- 패키지: `"+json.dumps(packages,ensure_ascii=False)+"`\n- 실행 명령: `python \""+str(script)+"\" --context \""+str(context)+"\" --output-dir \""+str(output)+"\" --version "+version+"`\n\n## 핵심 지표 DB 재계산\n\n"+"\n".join(core_lines)+"\n\n## 완료 범위\n\n- A1·A2·B1·B2·B3·세션 4개 결과 세그먼트·H3: 완료\n- 모든 assertion: PASS (`"+str(len(assertions))+"`건)\n- Excel 저장 후 재열기 검사: PASS\n- 입력 파일 해시·크기·수정시각 불변: PASS\n\n## 해석 제한\n\n- 원본 검정은 296검색·43세션의 탐색적 분석이다.\n- 합성 검정은 생성 구조·방향 보존과 코드 작동을 확인하는 모형 내부 진단이다. 합성 p값을 실제 사용자 모집단의 가설 채택·기각 근거로 사용하지 않는다.\n- 검색 단위 Fisher 검정은 동일 세션 반복 검색에 따른 독립성 한계가 있다.\n- A1은 부분 채택이며 입력 상태 컬럼이 없어 조건 제한 효과와 입력 품질 효과를 분리하지 못했다. A2는 기술통계, B1은 관련성, B2는 탐색지속성 교란 가능성이 있는 비인과 진단, H3는 탐색적 관계다.\n- hotel_click만 상세진입 KPI로 사용했다. BOOKING은 성과 KPI나 가설검정에 쓰지 않았다.\n\n## 산출물 해시\n\n- 분석 코드: `"+str(script)+"` / `"+code_hash+"`\n- Excel: `"+str(xlsx)+"` / `"+excel_hash+"`\n- 실행기록: 이 파일의 SHA-256은 자기참조를 피하기 위해 파일 외부 최종 인수인계에서 기록한다.\n\n## 금지 작업\n\n생성 코드·config 미실행·미수정. S0 1,000명 재생성, 10,000명, STRESS, S1~S3, OVERSAMPLED, SX/A3, BOOKING·예약 이벤트 데이터 생성·분석을 수행하지 않았다.\n"
    log.write_text(log_text,encoding="utf-8")
    print(json.dumps({"STEP2":"PASS","context_sha256":sha256(context),"core":{d:results[d]["core_pairs"] for d in results},"artifacts":{"code":{"path":str(script),"sha256":code_hash},"excel":{"path":str(xlsx),"sha256":excel_hash},"log":{"path":str(log),"sha256":sha256(log)}},"assertions":len(assertions)},ensure_ascii=True,indent=2))


if __name__ == "__main__": main()
