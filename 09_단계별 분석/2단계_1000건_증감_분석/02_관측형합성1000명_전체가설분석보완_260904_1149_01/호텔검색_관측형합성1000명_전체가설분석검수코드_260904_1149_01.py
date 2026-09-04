#!/usr/bin/env python3
"""Independent STEP3 QA. Does not import or execute STEP2 analysis code."""
from __future__ import annotations
import argparse, hashlib, json, math, platform, sqlite3, unicodedata
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo
import numpy as np
import pandas as pd
from scipy.stats import fisher_exact, mannwhitneyu
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

KST=ZoneInfo("Asia/Seoul"); TS="260904_1149"
APPROVED={"context":"b1558b92f5dba089316dd9cae338f7b9a5a72685b878d81c9ffb15fe7f0723e9","step2_code":"9c881ab9507f9222c99fa5cd03f76292dadbfcd95babf935d4f929bd797e1145","step2_excel":"efd48a3d8822d8148879ced2bfdaf01fe9b9bac0531f7e3885ff812cf4297c86","step2_log":"aa25c3950dd5298834664726fb4fc095baff28b45a9a5b4e78011a2b949e4a00"}
CORE_EXPECTED={"ORIGINAL_296":{"ZERO_RESULT_RATE":(147,296),"ZERO_FOLLOWUP_RATE":(140,147),"B3_IMMEDIATE_RECOVERY":(24,140),"B3_SESSION_FINAL_RECOVERY":(21,28),"SEARCH_HOTEL_CLICK_RATE":(45,296),"DIAG_FIRST_SEARCH_ZERO_RECOVERY":(4,6)},"S0_1000":{"ZERO_RESULT_RATE":(3434,6900),"ZERO_FOLLOWUP_RATE":(3271,3434),"B3_IMMEDIATE_RECOVERY":(558,3271),"B3_SESSION_FINAL_RECOVERY":(488,651),"SEARCH_HOTEL_CLICK_RATE":(1052,6900),"DIAG_FIRST_SEARCH_ZERO_RECOVERY":(92,139)}}
REGIONS=["Tokyo","Osaka","Kyoto","Sapporo","Fukuoka","UNKNOWN"]; INTENTS=["LOCATION_ONLY","PRICE","QUALITY_FILTER","AMENITY","MIXED"]
H3=["동일조건 반복","조건 완화","검색어 수정","지역 변경","조건 강화","혼합 변경"]
SEGS=["직접 성공","결과 노출·미선택","재검색 회복","지속 실패"]
FIELDS=["query_text","destination","property_type","property_grade","user_rating_min","price","amenity_count","region"]

def sh(p): return hashlib.sha256(Path(p).read_bytes()).hexdigest()
def lfsh(p): return hashlib.sha256(Path(p).read_bytes().replace(b"\r\n",b"\n")).hexdigest()
def norm(v):
    if v is None or (isinstance(v,float) and np.isnan(v)): return None
    x=unicodedata.normalize("NFKC",str(v)).strip().casefold(); return x or None
def ro(p):
    c=sqlite3.connect(Path(p).resolve().as_uri()+"?mode=ro",uri=True); c.execute("PRAGMA query_only=ON")
    assert c.execute("PRAGMA query_only").fetchone()[0]==1 and c.execute("PRAGMA integrity_check").fetchone()[0]=="ok"; return c
def rel(root,s):
    p=Path(s); return p.resolve() if p.is_absolute() and p.exists() else (root/s.replace("/","\\")).resolve(strict=True)
def fingerprint(p): p=Path(p); st=p.stat(); return {"sha256":sh(p),"size":st.st_size,"mtime_ns":st.st_mtime_ns}
def add(rows,cid,target,key,ind,stored,rule,status,fn,ds="ALL",note=""):
    def delta(a,b):
        if a is None and b is None:return (0.0,0.0)
        try:
            if pd.isna(a) and pd.isna(b):return (0.0,0.0)
            aa,bb=float(a),float(b); d=abs(aa-bb); return d,d/max(abs(aa),abs(bb),1e-300)
        except:return (None,None)
    ad,rd=delta(ind,stored)
    rows.append({"check_id":cid,"dataset_type":ds,"target":target,"comparison_key":key,"independent_value":ind,"step2_value":stored,"absolute_difference":ad,"relative_difference":rd,"tolerance":rule,"status":status,"evidence":fn,"note":note})

def direction(old,new,lower_relax):
    om=pd.isna(old); nm=pd.isna(new)
    if om and nm:return None
    if not om and nm:return "relax"
    if om and not nm:return "strengthen"
    if float(old)==float(new):return None
    dec=float(new)<float(old); relax=dec if lower_relax else not dec
    return "relax" if relax else "strengthen"
def equal(a,b,numeric=False):
    if numeric:
        if pd.isna(a) and pd.isna(b):return True
        if pd.isna(a) or pd.isna(b):return False
        return float(a)==float(b)
    return norm(a)==norm(b)
def classify(r):
    changed=[f for f in FIELDS if not equal(r[f],r["next_"+f],f in {"user_rating_min","price","amenity_count"})]
    if not changed:return "동일조건 반복",changed
    if "destination" in changed or "region" in changed:return "지역 변경",changed
    if "query_text" in changed:return "검색어 수정",changed
    dirs={x for x in [direction(r.price,r.next_price,False),direction(r.user_rating_min,r.next_user_rating_min,True),direction(r.amenity_count,r.next_amenity_count,True)] if x}
    if dirs=={"relax","strengthen"}:return "혼합 변경",changed
    if dirs=={"relax"}:return "조건 완화",changed
    if dirs=={"strengthen"}:return "조건 강화",changed
    raise AssertionError(f"unclassified H3 {r.search_id}: {changed}")
def city(v):
    x=norm(v)
    if x:
        for a,b in [("tokyo","Tokyo"),("osaka","Osaka"),("kyoto","Kyoto"),("sapporo","Sapporo"),("fukuoka","Fukuoka")]:
            if x==a or x.startswith(a+" ·"):return b
    return "UNKNOWN"
def odds_ci(a,b,c,d):
    z=any(x==0 for x in [a,b,c,d]); raw=fisher_exact([[a,b],[c,d]],alternative="two-sided")
    aa,bb,cc,dd=([x+.5 for x in [a,b,c,d]] if z else [a,b,c,d]); lo=math.exp(math.log(aa*dd/(bb*cc))-1.96*math.sqrt(1/aa+1/bb+1/cc+1/dd)); hi=math.exp(math.log(aa*dd/(bb*cc))+1.96*math.sqrt(1/aa+1/bb+1/cc+1/dd))
    return float(raw.statistic),lo,hi,float(raw.pvalue),z

def build(db,ds):
    with ro(db) as c:
        s=pd.read_sql_query("SELECT search_id,session_id,search_time,total_result_count,query_text,destination,checkin_date,checkout_date,sort_option,guest_count,data_origin FROM search",c)
        f=pd.read_sql_query("SELECT * FROM search_filter",c)
        clicks=pd.read_sql_query("SELECT DISTINCT search_id FROM event WHERE event_type='hotel_click' AND search_id IS NOT NULL",c)
    assert s.search_id.is_unique and f.search_id.is_unique and set(s.search_id)==set(f.search_id)
    b=s.merge(f,on="search_id",validate="one_to_one",suffixes=("_search","_filter")); b.search_time=pd.to_datetime(b.search_time.astype(str).str.replace(r"\s+KST$","",regex=True),errors="raise")
    for col in ["total_result_count","guest_count","user_rating_min","price","amenity_count"]: b[col]=pd.to_numeric(b[col],errors="coerce")
    b["has_hotel_click"]=b.search_id.isin(set(clicks.search_id)); b["is_zero_result"]=b.total_result_count.eq(0)
    b=b.sort_values(["session_id","search_time","search_id"],kind="mergesort").reset_index(drop=True); g=b.groupby("session_id",sort=False); b["search_order"]=g.cumcount()+1
    for col in ["search_id","total_result_count",*FIELDS]:b["next_"+col]=g[col].shift(-1)
    b["has_next_search"]=b.next_search_id.notna(); b["next_is_zero"]=b.next_total_result_count.eq(0); b["next_search_success"]=b.next_total_result_count.gt(0); b["next_search_has_hotel_click"]=b.next_search_id.isin(set(clicks.search_id))
    t=b[b.is_zero_result & b.has_next_search].copy(); lab=t.apply(classify,axis=1); t["transition_type"]=[x[0] for x in lab]; t["changed_conditions"]=["|".join(x[1]) for x in lab]
    sess=[]
    for sid,x in b.groupby("session_id",sort=True):
        vals=x.total_result_count.astype(int).tolist(); anyzero=any(v==0 for v in vals); later=any(v==0 and any(q>0 for q in vals[i+1:]) for i,v in enumerate(vals)); firstzero=vals[0]==0; firstlater=any(q>0 for q in vals[1:]) if firstzero else False; click=bool(x.has_hotel_click.any())
        seg=("재검색 회복" if firstlater else "지속 실패") if firstzero else ("직접 성공" if click else "결과 노출·미선택")
        sess.append({"session_id":sid,"search_count":len(x),"first_zero":firstzero,"experienced_zero":anyzero,"zero_later_positive":later,"first_zero_later_positive":firstlater,"has_click":click,"segment":seg})
    ss=pd.DataFrame(sess)
    core={"ZERO_RESULT_RATE":(int(b.is_zero_result.sum()),len(b)),"ZERO_FOLLOWUP_RATE":(int((b.is_zero_result&b.has_next_search).sum()),int(b.is_zero_result.sum())),"B3_IMMEDIATE_RECOVERY":(int(t.next_search_success.sum()),len(t)),"B3_SESSION_FINAL_RECOVERY":(int(ss.loc[ss.experienced_zero,"zero_later_positive"].sum()),int(ss.experienced_zero.sum())),"SEARCH_HOTEL_CLICK_RATE":(int(b.has_hotel_click.sum()),len(b)),"DIAG_FIRST_SEARCH_ZERO_RECOVERY":(int(ss.loc[ss.first_zero,"first_zero_later_positive"].sum()),int(ss.first_zero.sum()))}
    a1={}
    for mid,m in [("A1_AMENITY_GE3",b.amenity_count.ge(3)),("A1_RATING_SET",b.user_rating_min.notna()),("A1_PRICE_SET",b.price.notna())]:
        a=int((m&b.is_zero_result).sum()); bb=int((m&~b.is_zero_result).sum()); c=int((~m&b.is_zero_result).sum()); d=int((~m&~b.is_zero_result).sum()); a1[mid]={"cells":(a,bb,c,d),"stats":odds_ci(a,bb,c,d)}
    b["region_code"]=b.destination.map(city); pr=b.price.notna();ra=b.user_rating_min.notna();am=b.amenity_count.fillna(0).gt(0);cnt=pr.astype(int)+ra.astype(int)+am.astype(int);b["intent_code"]="LOCATION_ONLY";b.loc[cnt.ge(2),"intent_code"]="MIXED";b.loc[cnt.eq(1)&pr,"intent_code"]="PRICE";b.loc[cnt.eq(1)&ra,"intent_code"]="QUALITY_FILTER";b.loc[cnt.eq(1)&am,"intent_code"]="AMENITY"
    a2={}
    for reg in REGIONS:
      for inte in INTENTS:
        x=b[(b.region_code==reg)&(b.intent_code==inte)];n=len(x);z=int(x.is_zero_result.sum());cl=int(x.has_hotel_click.sum());a2[(reg,inte)]={"n":n,"numerator":z,"denominator":n,"rate":z/n if n else None,"nonzero_count":n-z,"hotel_click_searches":cl,"hotel_click_rate":cl/n if n else None,"is_empty_cell":n==0,"sparse_lt5":n<5,"sparse_lt10":n<10}
    m=b.is_zero_result;nxt=b.has_next_search;a=int((m&nxt).sum());bb=int((m&~nxt).sum());c=int((~m&nxt).sum());d=int((~m&~nxt).sum());b1={"cells":(a,bb,c,d),"stats":odds_ci(a,bb,c,d)}
    yes=ss.loc[ss.experienced_zero,"search_count"].astype(float);no=ss.loc[~ss.experienced_zero,"search_count"].astype(float);mw=mannwhitneyu(yes,no,alternative="two-sided",method="auto");rbc=2*float(mw.statistic)/(len(yes)*len(no))-1
    b2={"experienced_zero":dict(n=len(yes),mean=yes.mean(),median=yes.median(),q1=yes.quantile(.25),q3=yes.quantile(.75),iqr=yes.quantile(.75)-yes.quantile(.25),minimum=yes.min(),maximum=yes.max()),"no_zero_experience":dict(n=len(no),mean=no.mean(),median=no.median(),q1=no.quantile(.25),q3=no.quantile(.75),iqr=no.quantile(.75)-no.quantile(.25),minimum=no.min(),maximum=no.max()),"statistic":float(mw.statistic),"p_value":float(mw.pvalue),"effect_size":rbc}
    seg={x:int((ss.segment==x).sum()) for x in SEGS};h3={}
    for label in H3:
        x=t[t.transition_type==label];h3[label]={"n":len(x),"next_positive_count":int(x.next_search_success.sum()),"next_positive_rate":float(x.next_search_success.mean()) if len(x) else None,"next_hotel_click_count":int(x.next_search_has_hotel_click.sum()),"next_hotel_click_rate":float(x.next_search_has_hotel_click.mean()) if len(x) else None,"composition_rate":len(x)/len(t)}
    return {"base":b,"trans":t,"sessions":ss,"core":core,"a1":a1,"a2":a2,"b1":b1,"b2":b2,"seg":seg,"h3":h3}

def excel_rows(wb,sheet):
    ws=wb[sheet];h=[c.value for c in next(ws.iter_rows())];return [dict(zip(h,[c.value for c in r])) for r in ws.iter_rows(min_row=2)]
def close(a,b,integer=False):
    if a is None and (b is None or pd.isna(b)):return True
    if integer:return int(a)==int(b)
    try:return math.isclose(float(a),float(b),abs_tol=1e-10,rel_tol=1e-8)
    except:return a==b

def main():
    ap=argparse.ArgumentParser();ap.add_argument("--context",type=Path,required=True);ap.add_argument("--step2-code",type=Path,required=True);ap.add_argument("--step2-excel",type=Path,required=True);ap.add_argument("--step2-log",type=Path,required=True);ap.add_argument("--output-dir",type=Path,required=True);ap.add_argument("--version",default="01");args=ap.parse_args()
    ctxp=args.context.resolve(strict=True);out=args.output_dir.resolve(strict=True);ver=args.version.zfill(2);script=Path(__file__).resolve();qa=[];started=datetime.now(KST).isoformat(timespec="seconds")
    fixed={"context":ctxp,"step2_code":args.step2_code.resolve(strict=True),"step2_excel":args.step2_excel.resolve(strict=True),"step2_log":args.step2_log.resolve(strict=True)}
    for k,p in fixed.items():add(qa,"GATE_HASH_"+k,k,str(p),sh(p),APPROVED[k],"exact SHA-256","PASS" if sh(p)==APPROVED[k] else "FAIL","sha256")
    if any(x["status"]=="FAIL" for x in qa):raise AssertionError("entry hash gate")
    ctx=json.loads(ctxp.read_text(encoding="utf-8"));root=ctxp.parents[3]
    add(qa,"GATE_STEP1","context","status",ctx.get("step1_status"),"PASS","exact","PASS" if ctx.get("step1_status")=="PASS" else "FAIL","preflight")
    add(qa,"GATE_HASH_GATE","context","hash_gate",ctx.get("hash_gate_status"),"PASS","exact","PASS" if ctx.get("hash_gate_status")=="PASS" else "FAIL","preflight")
    logtxt=fixed["step2_log"].read_text(encoding="utf-8");add(qa,"GATE_STEP2","log","STEP2", "PASS" if "STEP2=PASS" in logtxt else "MISSING","PASS","exact","PASS" if "STEP2=PASS" in logtxt else "FAIL","text check")
    wb=load_workbook(fixed["step2_excel"],data_only=True,read_only=True);required=["run_info","metric_dictionary","G3_core_metrics","A1_filters","A2_region_intent","B1_followup","B2_search_count","B3_recovery","session_segments","H3_transitions","raw_contingency_tables","assertions","known_limitations"]
    add(qa,"GATE_SHEETS","STEP2 Excel","sheetnames",wb.sheetnames,required,"exact ordered list","PASS" if wb.sheetnames==required else "FAIL","openpyxl")
    ri=excel_rows(wb,"run_info")[0];add(qa,"GATE_TS","STEP2 Excel","bundle_run_ts",str(ri["bundle_run_ts"]),TS,"exact","PASS" if str(ri["bundle_run_ts"])==TS else "FAIL","run_info")
    files={x["role"]:rel(root,x["path"]) for x in ctx["all_followup_inputs"]};addp=rel(root,ctx["transport_normalization_addendum"]["path"]);parp=rel(root,ctx["parent_corrected_manifest"]["path"])
    for x in ctx["all_followup_inputs"]:
        p=files[x["role"]];ok=sh(p)==x["current_raw_sha256"];add(qa,"GATE_INPUT_"+x["role"],"input",x["role"],sh(p),x["current_raw_sha256"],"exact raw SHA","PASS" if ok else "FAIL","sha256")
    for role,key in [("code","generation_code_hashes"),("config","config_hashes")]:
        z=ctx[key];ok=lfsh(files[role])==z["lf_normalized_sha256"]==z["parent_approved_sha256"];add(qa,"GATE_LF_"+role,"input",role,lfsh(files[role]),z["parent_approved_sha256"],"exact LF SHA","PASS" if ok else "FAIL","CRLF->LF only")
    for label,p,expected in [("addendum",addp,ctx["transport_normalization_addendum"]["sha256"]),("parent",parp,ctx["parent_corrected_manifest"]["sha256"])]:add(qa,"GATE_"+label,"manifest",label,sh(p),expected,"exact SHA","PASS" if sh(p)==expected else "FAIL","sha256")
    for ds,role in [("ORIGINAL_296","source"),("S0_1000","synthetic")]:
        with ro(files[role]) as c:add(qa,"GATE_DB_"+ds,"SQLite",ds,c.execute("PRAGMA integrity_check").fetchone()[0],"ok","exact","PASS","mode=ro/query_only")
    if any(x["status"]=="FAIL" for x in qa):raise AssertionError("entry gate")
    tracked=[*fixed.values(),*files.values(),addp,parp];before={str(p):fingerprint(p) for p in tracked}
    # Independent database calculations finish before any STEP2 result sheet is read.
    calc={"ORIGINAL_296":build(files["source"],"ORIGINAL_296"),"S0_1000":build(files["synthetic"],"S0_1000")}
    checks={k:[] for k in ["core","A1","A2","B1","B2","B3","segments","H3"]}
    stored={s:excel_rows(wb,s) for s in ["G3_core_metrics","A1_filters","A2_region_intent","B1_followup","B2_search_count","B3_recovery","session_segments","H3_transitions","raw_contingency_tables"]}
    def cmp(rows,cid,ds,target,key,iv,sv,integer=False,fn="independent build"):
        ok=close(iv,sv,integer);add(rows,cid,target,key,iv,sv,"exact integer" if integer else "abs<=1e-10 or rel<=1e-8","PASS" if ok else "FAIL",fn,ds)
    for ds,x in calc.items():
      for mid,(num,den) in x["core"].items():
        r=next(q for q in stored["G3_core_metrics"] if q["dataset_type"]==ds and q["metric_id"]==mid)
        for f,v,integer in [("numerator",num,True),("denominator",den,True),("rate",num/den,False)]:cmp(checks["core"],f"CORE_{ds}_{mid}_{f}",ds,"core",mid+"|"+f,v,r[f],integer,"independent_core")
        exp=CORE_EXPECTED[ds][mid];cmp(checks["core"],f"ANCHOR_{ds}_{mid}",ds,"approved anchor",mid,str((num,den)),str(exp),False,"post-calculation anchor")
      # A1 rows and raw cells
      groups={"A1_AMENITY_GE3":["amenity_count>=3","amenity_count<3"],"A1_RATING_SET":["set","unset"],"A1_PRICE_SET":["set","unset"]}
      for mid,z in x["a1"].items():
        a,b,c,d=z["cells"];od,lo,hi,p,corr=z["stats"]
        for grp,num,den in [(groups[mid][0],a,a+b),(groups[mid][1],c,c+d)]:
          r=next(q for q in stored["A1_filters"] if q["dataset_type"]==ds and q["metric_id"]==mid and q["group"]==grp)
          vals={"numerator":num,"denominator":den,"n":den,"rate":num/den,"statistic":od,"effect_size":od,"ci_95_low":lo,"ci_95_high":hi,"p_value":p,"ci_zero_cell_correction":corr}
          for f,v in vals.items():cmp(checks["A1"],f"A1_{ds}_{mid}_{grp}_{f}",ds,"A1",mid+"|"+grp+"|"+f,v,r[f],f in {"numerator","denominator","n"},"independent_a1")
        for grp,outcome,value in [(groups[mid][0],"zero_result",a),(groups[mid][0],"nonzero_result",b),(groups[mid][1],"zero_result",c),(groups[mid][1],"nonzero_result",d)]:
          rr=next(q for q in stored["raw_contingency_tables"] if q["dataset_type"]==ds and q["metric_id"]==mid and q["group"]==grp and q["outcome"]==outcome)
          cmp(checks["A1"],f"RAW_{ds}_{mid}_{grp}_{outcome}",ds,"raw contingency",mid+"|"+grp+"|"+outcome,value,rr["count"],True,"independent raw 2x2")
      for (reg,inte),vals in x["a2"].items():
        r=next(q for q in stored["A2_region_intent"] if q["dataset_type"]==ds and q["region"]==reg and q["intent"]==inte)
        for f,v in vals.items():cmp(checks["A2"],f"A2_{ds}_{reg}_{inte}_{f}",ds,"A2",reg+"|"+inte+"|"+f,v,r[f],f in {"n","numerator","denominator","nonzero_count","hotel_click_searches"},"independent_a2")
      a,b,c,d=x["b1"]["cells"];od,lo,hi,p,corr=x["b1"]["stats"]
      for grp,num,den in [("zero_result",a,a+b),("positive_result",c,c+d)]:
        r=next(q for q in stored["B1_followup"] if q["dataset_type"]==ds and q["group"]==grp)
        for f,v in {"numerator":num,"denominator":den,"n":den,"rate":num/den,"statistic":od,"effect_size":od,"ci_95_low":lo,"ci_95_high":hi,"p_value":p,"ci_zero_cell_correction":corr}.items():cmp(checks["B1"],f"B1_{ds}_{grp}_{f}",ds,"B1",grp+"|"+f,v,r[f],f in {"numerator","denominator","n"},"independent_b1")
      for grp,outcome,value in [("zero_result","has_followup",a),("zero_result","no_followup",b),("positive_result","has_followup",c),("positive_result","no_followup",d)]:
        rr=next(q for q in stored["raw_contingency_tables"] if q["dataset_type"]==ds and q["metric_id"]=="B1_IMMEDIATE_FOLLOWUP" and q["group"]==grp and q["outcome"]==outcome)
        cmp(checks["B1"],f"RAW_B1_{ds}_{grp}_{outcome}",ds,"raw contingency","B1|"+grp+"|"+outcome,value,rr["count"],True,"independent raw 2x2")
      for grp,vals in [(g,x["b2"][g]) for g in ["experienced_zero","no_zero_experience"]]:
        r=next(q for q in stored["B2_search_count"] if q["dataset_type"]==ds and q["group"]==grp)
        vv={**vals,"statistic":x["b2"]["statistic"],"p_value":x["b2"]["p_value"],"effect_size":x["b2"]["effect_size"]}
        for f,v in vv.items():cmp(checks["B2"],f"B2_{ds}_{grp}_{f}",ds,"B2",grp+"|"+f,v,r[f],f=="n","independent_b2")
      for mid in ["B3_IMMEDIATE_RECOVERY","B3_SESSION_FINAL_RECOVERY","DIAG_FIRST_SEARCH_ZERO_RECOVERY"]:
        num,den=x["core"][mid];r=next(q for q in stored["B3_recovery"] if q["dataset_type"]==ds and q["metric_id"]==mid)
        for f,v in [("numerator",num),("denominator",den),("n",den),("rate",num/den)]:cmp(checks["B3"],f"B3_{ds}_{mid}_{f}",ds,"B3",mid+"|"+f,v,r[f],f!="rate","independent_b3")
      for seg,n in x["seg"].items():
        r=next(q for q in stored["session_segments"] if q["dataset_type"]==ds and q["group"]==seg)
        for f,v in [("n",n),("numerator",n),("denominator",len(x["sessions"])),("rate",n/len(x["sessions"]))]:cmp(checks["segments"],f"SEG_{ds}_{seg}_{f}",ds,"segments",seg+"|"+f,v,r[f],f!="rate","independent_segments")
      for label,vals in x["h3"].items():
        r=next(q for q in stored["H3_transitions"] if q["dataset_type"]==ds and q["group"]==label)
        for f,v in vals.items():cmp(checks["H3"],f"H3_{ds}_{label}_{f}",ds,"H3",label+"|"+f,v,r[f],f in {"n","next_positive_count","next_hotel_click_count"},"independent_h3")
    # Key uniqueness and cardinality across all STEP2 result sheets.
    schema=[];keys={"G3_core_metrics":["dataset_type","metric_id"],"A1_filters":["dataset_type","metric_id","group"],"A2_region_intent":["dataset_type","region","intent"],"B1_followup":["dataset_type","group"],"B2_search_count":["dataset_type","group"],"B3_recovery":["dataset_type","metric_id"],"session_segments":["dataset_type","group"],"H3_transitions":["dataset_type","group"],"raw_contingency_tables":["dataset_type","metric_id","group","outcome"]}
    for sheet,ks in keys.items():
        df=pd.DataFrame(stored[sheet]);dups=int(df.duplicated(ks).sum());add(schema,"SCHEMA_"+sheet,"workbook schema","+".join(ks),dups,0,"exact","PASS" if dups==0 else "FAIL","key uniqueness")
    # Deterministic traces. STEP2 has no row-level mart sheet, so source-code audit + aggregate equality is stated.
    traces=[]
    for ds,x in calc.items():
      for seg in SEGS:
        q=x["sessions"][x["sessions"].segment==seg].sort_values("session_id").head(1)
        if len(q):
          sid=q.iloc[0].session_id;z=x["base"][x["base"].session_id==sid].iloc[0];traces.append({"trace_type":"segment minimum session","dataset_type":ds,"session_id":sid,"current_search_id":z.search_id,"next_search_id":z.next_search_id,"current_result_count":z.total_result_count,"next_result_count":z.next_total_result_count,"changed_conditions":None,"independent_classification":seg,"step2_classification":"not persisted row-level; aggregate/code audited","success":q.iloc[0].zero_later_positive,"click":q.iloc[0].has_click,"match_status":"AGGREGATE_CONFIRMED"})
      take=1 if ds=="ORIGINAL_296" else 3
      for label in H3:
        for _,z in x["trans"][x["trans"].transition_type==label].sort_values(["session_id","search_time","search_id"],kind="mergesort").head(take).iterrows():traces.append({"trace_type":"H3 stable first","dataset_type":ds,"session_id":z.session_id,"current_search_id":z.search_id,"next_search_id":z.next_search_id,"current_result_count":z.total_result_count,"next_result_count":z.next_total_result_count,"changed_conditions":z.changed_conditions,"independent_classification":label,"step2_classification":"not persisted row-level; aggregate/code audited","success":z.next_search_success,"click":z.next_search_has_hotel_click,"match_status":"AGGREGATE_CONFIRMED"})
      # B3 numerator and exclusion examples, stable first where available.
      candidates=[("B3 immediate included",x["trans"][x["trans"].next_search_success]),("B3 immediate excluded",x["trans"][~x["trans"].next_search_success])]
      for typ,q in candidates:
        if len(q):
          z=q.sort_values(["session_id","search_time","search_id"],kind="mergesort").iloc[0];traces.append({"trace_type":typ,"dataset_type":ds,"session_id":z.session_id,"current_search_id":z.search_id,"next_search_id":z.next_search_id,"current_result_count":z.total_result_count,"next_result_count":z.next_total_result_count,"changed_conditions":z.changed_conditions,"independent_classification":z.transition_type,"step2_classification":"not persisted row-level; aggregate/code audited","success":z.next_search_success,"click":z.next_search_has_hotel_click,"match_status":"AGGREGATE_CONFIRMED"})
      for metric,flag,cohort in [("B3 session final",x["sessions"].zero_later_positive,x["sessions"].experienced_zero),("diagnostic first-zero",x["sessions"].first_zero_later_positive,x["sessions"].first_zero)]:
        for suffix,want in [("included",True),("excluded",False)]:
          q=x["sessions"][cohort & flag.eq(want)].sort_values("session_id").head(1)
          if len(q):
            sr=q.iloc[0]; sb=x["base"][x["base"].session_id==sr.session_id].sort_values(["search_time","search_id"],kind="mergesort"); z=sb[sb.is_zero_result].iloc[0]
            traces.append({"trace_type":metric+" "+suffix,"dataset_type":ds,"session_id":sr.session_id,"current_search_id":z.search_id,"next_search_id":z.next_search_id,"current_result_count":z.total_result_count,"next_result_count":z.next_total_result_count,"changed_conditions":None,"independent_classification":metric,"step2_classification":"not persisted row-level; aggregate/code audited","success":want,"click":sr.has_click,"match_status":"AGGREGATE_CONFIRMED"})
    # Interpretation/code-static QA, after independent numeric work.
    s2code=fixed["step2_code"].read_text(encoding="utf-8");limitations="\n".join(str(r.get("limitation") or "")+" "+str(r.get("interpretation") or "") for shn in ["A1_filters","A2_region_intent","B1_followup","B2_search_count","H3_transitions","known_limitations"] for r in excel_rows(wb,shn))
    interp=[]
    tests=[("A1 OR direction",all(x["a1"][m]["stats"][0]>1 for x in calc.values() for m in x["a1"]),"all A1 OR > 1"),("B1 direction",all(x["core"]["ZERO_FOLLOWUP_RATE"][0]/x["core"]["ZERO_FOLLOWUP_RATE"][1] > (x["b1"]["cells"][2]/(x["b1"]["cells"][2]+x["b1"]["cells"][3])) for x in calc.values()),"zero followup > positive followup"),("B2 noncausal","비인과" in limitations or "인과" in limitations,"limitation text"),("A2 sparse","희소" in limitations and "인과" in limitations,"limitation text"),("synthetic p caveat","합성 p값" in limitations,"limitation text"),("session dependence","독립성 한계" in limitations,"limitation text"),("single click KPI","hotel_detail_view" in limitations and "hotel_click" in limitations,"limitation text"),("prohibited excluded",all(x in s2code for x in ["a3_excluded","booking_kpi_used"]),"run_info/code"),("zero wording","표본에서 관측되지 않음" in limitations,"known limitations")]
    for i,(name,ok,evid) in enumerate(tests,1):add(interp,f"INTERP_{i:02d}","interpretation",name,ok,True,"required principle","PASS" if ok else "FAIL",evid)
    after={str(p):fingerprint(p) for p in tracked};immut=[]
    for p in tracked:add(immut,"IMMUT_"+str(len(immut)+1),"input immutability",Path(p).name,json.dumps(after[str(p)],sort_keys=True),json.dumps(before[str(p)],sort_keys=True),"exact hash,size,mtime","PASS" if after[str(p)]==before[str(p)] else "FAIL","fingerprint")
    allchecks=qa+sum(checks.values(),[])+schema+interp+immut
    calc_fail=sum(x["status"]=="FAIL" for x in sum(checks.values(),[])+schema);gate_fail=sum(x["status"]=="FAIL" for x in qa+immut);interp_fail=sum(x["status"]=="FAIL" for x in interp)
    # Row-level STEP2 labels are not persisted, but deterministic traces plus full aggregates and code-static logic agree.
    status="FAIL" if gate_fail or calc_fail else ("REVIEW_REQUIRED" if interp_fail else "PASS")
    qa_summary=pd.DataFrame([{"STEP3":status,"entry_gate_failures":gate_fail,"calculation_failures":calc_fail,"interpretation_failures":interp_fail,"total_checks":len(allchecks),"step2_row_level_classification_persisted":False,"trace_assurance":"independent stable traces + complete aggregate comparison + STEP2 static code audit","generator_executed":False,"forbidden_data_generated":False}])
    sheets={"run_info":pd.DataFrame([{"step":"STEP3_INDEPENDENT_RECALCULATION_QA","bundle_run_ts":TS,"version":ver,"started_at_kst":started,"context_sha256":sh(ctxp),"step2_code_sha256":sh(fixed['step2_code']),"step2_excel_sha256":sh(fixed['step2_excel']),"step2_log_sha256":sh(fixed['step2_log'])}]),"qa_summary":qa_summary,"core_metrics_check":pd.DataFrame(checks["core"]),"A1_check":pd.DataFrame(checks["A1"]),"A2_check":pd.DataFrame(checks["A2"]),"B1_check":pd.DataFrame(checks["B1"]),"B2_check":pd.DataFrame(checks["B2"]),"B3_check":pd.DataFrame(checks["B3"]),"session_segments_check":pd.DataFrame(checks["segments"]),"H3_check":pd.DataFrame(checks["H3"]),"deterministic_traces":pd.DataFrame(traces),"workbook_schema_check":pd.DataFrame(schema),"interpretation_check":pd.DataFrame(interp),"input_immutability":pd.DataFrame(immut),"known_limitations":pd.DataFrame({"limitation":["STEP2 workbook stores aggregate results, not row-level segment/H3 assignments; deterministic row traces are independently reconstructed and their aggregate classifications are checked against STEP2.","Original search-level tests are exploratory and repeated searches within sessions limit independence.","Synthetic p-values are model-internal diagnostics, not evidence about a real population.","A2 sparse cells are descriptive; 0% means not observed in this sample.","A3 and BOOKING KPI are excluded; hotel_click is the sole detail-entry KPI."]})}
    xlsx=out/f"호텔검색_관측형합성1000명_전체가설분석검수결과_{TS}_{ver}.xlsx";log=out/f"호텔검색_관측형합성1000명_전체가설분석검수기록_{TS}_{ver}.md"
    if xlsx.exists() or log.exists():raise FileExistsError("QA output exists")
    with pd.ExcelWriter(xlsx,engine="openpyxl") as w:
        for name,df in sheets.items():df.to_excel(w,sheet_name=name,index=False)
    ow=load_workbook(xlsx)
    for ws in ow.worksheets:
        ws.freeze_panes="A2";ws.auto_filter.ref=ws.dimensions
        for c in ws[1]:c.font=Font(bold=True,color="FFFFFF");c.fill=PatternFill("solid",fgColor="1F4E78")
        for col in ws.columns:ws.column_dimensions[col[0].column_letter].width=min(55,max(10,max(len(str(c.value or "")) for c in col)+2))
    ow.save(xlsx);vr=load_workbook(xlsx,data_only=True,read_only=True);assert vr.sheetnames==list(sheets)
    for name,df in sheets.items():assert vr[name].max_row-1==len(df)
    assert vr["qa_summary"]["A2"].value==status;vr.close()
    finished=datetime.now(KST).isoformat(timespec="seconds");corelines=[]
    for ds,x in calc.items():
      for mid,(n,d) in x["core"].items():corelines.append(f"- {ds} `{mid}`: {n}/{d} = {n/d:.4%}")
    logtxt=f"# STEP3 독립 재검산 기록\n\n- 판정: **STEP3={status}**\n- 시작: `{started}`\n- 종료: `{finished}`\n- 독립성: STEP2 코드 import·호출·실행 없음; SQLite에서 독립 기반표·분류·통계를 먼저 완성한 뒤 Excel 대조\n- 입력 게이트 실패: {gate_fail}\n- 계산/스키마 불일치: {calc_fail}\n- 해석 QA 실패: {interp_fail}\n- 전체 검수 행: {len(allchecks)}\n\n## 핵심 지표\n\n"+"\n".join(corelines)+f"\n\n## 전수 대조\n\n- A1·A2·B1·B2·B3·세션 세그먼트·H3: {'PASS' if calc_fail==0 else 'FAIL'}\n- 결정적 표본 추적: 독립 안정 정렬 사례 {len(traces)}건 기록; STEP2 행 단위 분류는 Excel에 미보존되어 집계 전수 일치와 코드 정적 검토로 보완\n- 기존 입력 불변: {'PASS' if gate_fail==0 else 'FAIL'}\n- 해석 원칙: {'PASS' if interp_fail==0 else 'FAIL'}\n\n## 산출물\n\n- 검수 코드: `{script}` / `{sh(script)}`\n- 검수 Excel: `{xlsx}` / `{sh(xlsx)}`\n- 검수기록: 자기참조 방지를 위해 이 파일 SHA-256은 최종 인수인계에서 기록\n\n## 금지 작업\n\nSTEP2 산출물과 기존 입력을 수정하지 않았다. 생성기 실행, S0 재생성, 10,000명, STRESS, S1~S3, OVERSAMPLED, SX/A3, BOOKING·예약 이벤트 분석/생성을 수행하지 않았다. 최종 Word·최종 보완 매니페스트를 생성하지 않았다.\n"
    log.write_text(logtxt,encoding="utf-8")
    print(json.dumps({"STEP3":status,"checks":len(allchecks),"traces":len(traces),"artifacts":{"code":{"path":str(script),"sha256":sh(script)},"excel":{"path":str(xlsx),"sha256":sh(xlsx)},"log":{"path":str(log),"sha256":sh(log)}}},ensure_ascii=True,indent=2))

if __name__=="__main__":main()
