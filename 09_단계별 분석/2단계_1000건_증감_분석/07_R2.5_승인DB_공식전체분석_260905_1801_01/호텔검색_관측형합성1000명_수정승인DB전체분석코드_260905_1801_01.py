#!/usr/bin/env python3
"""Run the approved STEP2 hypothesis logic against the R2 lineage-free database."""
from __future__ import annotations
import argparse, hashlib, importlib.metadata, importlib.util, json, math, platform, sqlite3, sys
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font,PatternFill
from scipy import stats
KST=ZoneInfo('Asia/Seoul');SEED=20260904;BOOT=500;APPROVED='7d449fab6847730be38fe46ebb6bd6a5b31690cdf38cf4bdbadeaa75edd1ae04'
def sha(p):
 h=hashlib.sha256()
 with Path(p).open('rb') as f:
  for b in iter(lambda:f.read(1<<20),b''):h.update(b)
 return h.hexdigest()
def ro(p):
 c=sqlite3.connect(Path(p).resolve().as_uri()+'?mode=ro',uri=True);c.row_factory=sqlite3.Row;c.execute('pragma query_only=on');assert c.execute('pragma query_only').fetchone()[0]==1;return c
def module(path,name):
 s=importlib.util.spec_from_file_location(name,path);m=importlib.util.module_from_spec(s);s.loader.exec_module(m);return m
def dt(v):return pd.to_datetime(pd.Series(v).astype(str).str.replace(r'\s+KST$','+09:00',regex=True),errors='coerce',utc=True)
def zero_timing(c,label):
 s=pd.read_sql_query('select session_id,search_id,search_time,total_result_count from search',c);s['t']=dt(s.search_time);g=[]
 for _,x in s.sort_values(['session_id','t','search_id'],kind='mergesort').groupby('session_id'):
  rr=list(x.itertuples());g += [(rr[i+1].t-r.t).total_seconds() for i,r in enumerate(rr[:-1]) if r.total_result_count==0]
 a=np.asarray(g,float);p=a[a>0];l=np.log(p);mu=float(l.mean());sigma=float(l.std(ddof=0));D=float(stats.kstest(p,'lognorm',args=(sigma,0,math.exp(mu))).statistic);rng=np.random.default_rng(SEED);sim=[]
 for _ in range(BOOT):
  z=rng.lognormal(mu,sigma,len(p));q=np.log(z);sim.append(stats.kstest(z,'lognorm',args=(q.std(ddof=0),0,math.exp(q.mean()))).statistic)
 return {'dataset_type':label,'transition_n':len(a),'negative_n':int((a<0).sum()),'zero_n':int((a==0).sum()),'positive_n':len(p),'mean_seconds':float(a.mean()),'median_seconds':float(np.median(a)),'q1_seconds':float(np.quantile(a,.25)),'q3_seconds':float(np.quantile(a,.75)),'p90_seconds':float(np.quantile(a,.9)),'p95_seconds':float(np.quantile(a,.95)),'max_seconds':float(a.max()),'lognormal_mu':mu,'lognormal_sigma':sigma,'ks_statistic':D,'bootstrap_p':(1+sum(x>=D for x in sim))/(BOOT+1),'status':'WARN'}
def cross(c,label):
 s=pd.read_sql_query('select search_id,session_id,search_time from search',c);e=pd.read_sql_query("select event_id,session_id,search_id,event_type,event_at from event where event_type not like 'booking_%'",c);s['t']=dt(s.search_time);e['t']=dt(e.event_at);nxt={}
 for _,g in s.sort_values(['session_id','t','search_id']).groupby('session_id'):
  rr=list(g.itertuples());nxt.update({x.search_id:(rr[i+1].t if i+1<len(rr) else None) for i,x in enumerate(rr)})
 cat={'before':0,'equal':0,'after':0,'no_next':0,'missing_key':0};over=[];ss=set();qq=set()
 for x in e.itertuples():
  if not x.search_id or x.search_id not in nxt or pd.isna(x.t):cat['missing_key']+=1
  elif nxt[x.search_id] is None:cat['no_next']+=1
  else:
   z=(x.t-nxt[x.search_id]).total_seconds();k='before' if z<0 else ('equal' if z==0 else 'after');cat[k]+=1
   if z>=0:over.append(z);ss.add(x.session_id);qq.add(x.search_id)
 n=cat['before']+cat['equal']+cat['after'];o=cat['equal']+cat['after'];return {'dataset_type':label,**cat,'comparable_n':n,'overlap_n':o,'overlap_rate':o/n,'affected_sessions':len(ss),'affected_searches':len(qq),'median_excess':float(np.median(over)),'q1_excess':float(np.quantile(over,.25)),'q3_excess':float(np.quantile(over,.75)),'p90_excess':float(np.quantile(over,.9)),'max_excess':max(over)}
def high_order(c,label):
 b=pd.read_sql_query('select s.total_result_count,f.* from search s join search_filter f using(search_id)',c);b['price_active']=b.price.notna();b['rating_active']=b.user_rating_min.notna();b['amenity_active']=b.amenity_count.fillna(0).gt(0);b['core_count']=b[['price_active','rating_active','amenity_active']].sum(axis=1);b['definition_a_count']=b.core_count+b.property_type.notna()+b.property_grade.notna()+b.region.notna();rows=[]
 for definition,col in [('B_core3','core_count'),('A_sensitivity6','definition_a_count')]:
  for n,g in b.groupby(col):rows.append({'dataset_type':label,'definition':definition,'active_count':int(n),'n':len(g),'zero_n':int((g.total_result_count==0).sum()),'zero_rate':float((g.total_result_count==0).mean())})
 return rows
def time_integrity(c):
 q=lambda s:c.execute(s).fetchone()[0]
 return [{'check':'search_internal_negative','failure_n':q("with x as(select datetime(replace(search_time,' KST','')) t,lag(datetime(replace(search_time,' KST',''))) over(partition by session_id order by search_time,search_id) p from search)select count(*) from x where t<p")},{'check':'event_internal_negative','failure_n':q("with x as(select datetime(replace(event_at,' KST','')) t,lag(datetime(replace(event_at,' KST',''))) over(partition by session_id order by event_at,event_id) p from event)select count(*) from x where t<p")},{'check':'end_before_last_event','failure_n':q("with x as(select session_id,max(datetime(replace(event_at,' KST',''))) a,max(datetime(replace(session_end_time,' KST',''))) z from event group by 1)select count(*) from x where z<a")},{'check':'end_before_last_search','failure_n':q("with e as(select session_id,max(datetime(replace(session_end_time,' KST',''))) z from event group by 1),s as(select session_id,max(datetime(replace(search_time,' KST',''))) a from search group by 1)select count(*) from e join s using(session_id)where z<a")}]
def main():
 if hasattr(sys.stdout,'reconfigure'):sys.stdout.reconfigure(encoding='utf-8')
 p=argparse.ArgumentParser();p.add_argument('--db',type=Path,required=True);p.add_argument('--source-db',type=Path,required=True);p.add_argument('--legacy-step2-code',type=Path,required=True);p.add_argument('--marts-dir',type=Path,required=True);p.add_argument('--lineage-json',type=Path,required=True);p.add_argument('--output-dir',type=Path,required=True);a=p.parse_args();out=a.output_dir.resolve();stamp='260905_1801_01';before=sha(a.db);assert before==APPROVED
 legacy=module(a.legacy_step2_code,'legacy_step2');sys.path.insert(0,str(a.marts_dir));import build_original_296_marts as marts
 start=datetime.now(KST).isoformat();res={'S0_TIME_ALIGNED_1000':legacy.calculate(a.db,'S0_TIME_ALIGNED_1000',marts)};c=ro(a.db);source=ro(a.source_db)
 timing=[zero_timing(source,'ORIGINAL_296'),zero_timing(c,'S0_TIME_ALIGNED_1000')];crosses=[cross(source,'ORIGINAL_296'),cross(c,'S0_TIME_ALIGNED_1000')];high=high_order(source,'ORIGINAL_296')+high_order(c,'S0_TIME_ALIGNED_1000');tint=time_integrity(c)
 lineage=json.loads(a.lineage_json.read_text(encoding='utf-8'));bn=[x for x in lineage['bn_definition_b'] if x['dataset'] in ['ORIGINAL_296','S0_TIME_ALIGNED']]
 tables={'run_info':pd.DataFrame([{'step':'STEP2_R2_5','input_db':str(a.db.resolve()),'input_sha256':before,'started_at_kst':start,'legacy_logic':str(a.legacy_step2_code.resolve()),'sqlite_access':'URI mode=ro; query_only=ON'}]),'metric_dictionary':legacy.metric_dictionary(),'G3_core_metrics':pd.DataFrame(res['S0_TIME_ALIGNED_1000']['core']),'A1_filters':pd.DataFrame(res['S0_TIME_ALIGNED_1000']['a1']),'A2_region_intent':pd.DataFrame(res['S0_TIME_ALIGNED_1000']['a2']),'B1_followup':pd.DataFrame(res['S0_TIME_ALIGNED_1000']['b1']),'B2_search_count':pd.DataFrame(res['S0_TIME_ALIGNED_1000']['b2']),'B3_recovery':pd.DataFrame(res['S0_TIME_ALIGNED_1000']['b3']),'session_segments':pd.DataFrame(res['S0_TIME_ALIGNED_1000']['segments']),'H3_transitions':pd.DataFrame(res['S0_TIME_ALIGNED_1000']['h3']),'raw_contingency':pd.DataFrame(res['S0_TIME_ALIGNED_1000']['raw']),'high_order':pd.DataFrame(high),'BN_definition_B':pd.DataFrame(bn),'zero_timing':pd.DataFrame(timing),'cross_stream':pd.DataFrame(crosses),'time_integrity':pd.DataFrame(tint)}
 x=out/f'호텔검색_관측형합성1000명_수정승인DB전체분석결과_{stamp}.xlsx';j=out/f'호텔검색_관측형합성1000명_수정승인DB전체분석결과_{stamp}.json'
 with pd.ExcelWriter(x,engine='openpyxl') as w:
  for n,d in tables.items():d.to_excel(w,sheet_name=n[:31],index=False)
 wb=load_workbook(x)
 for ws in wb:
  ws.freeze_panes='A2';ws.auto_filter.ref=ws.dimensions
  for z in ws[1]:z.font=Font(bold=True,color='FFFFFF');z.fill=PatternFill('solid',fgColor='1F4E78')
 wb.save(x)
 payload={'step':'STEP2','status':'PASS','input_sha256_before':before,'input_sha256_after':sha(a.db),'core_pairs':res['S0_TIME_ALIGNED_1000']['core_pairs'],'tables':{k:v.astype(object).where(pd.notna(v),None).to_dict('records') for k,v in tables.items()},'environment':{'python':platform.python_version(),'sqlite':sqlite3.sqlite_version,**{z:importlib.metadata.version(z) for z in ['pandas','numpy','scipy','openpyxl','matplotlib']}}};j.write_text(json.dumps(payload,ensure_ascii=False,indent=2,default=str),encoding='utf-8')
 fig,ax=plt.subplots(1,2,figsize=(11,4));hb=pd.DataFrame(high);z=hb[(hb.definition=='B_core3')&(hb.active_count==3)];ax[0].bar(z.dataset_type,z.zero_rate);ax[0].set_title('Definition B: 3 active filters');bb=pd.DataFrame(bn);ax[1].bar(bb.dataset,bb.roc_auc);ax[1].set_title('Constrained BN ROC-AUC');fig.tight_layout();fig.savefig(out/f'호텔검색_관측형합성1000명_고차결합BN시각화_{stamp}.png',dpi=160);plt.close(fig)
 fig,ax=plt.subplots(1,2,figsize=(11,4));tt=pd.DataFrame(timing);ax[0].bar(tt.dataset_type,tt.median_seconds);ax[0].set_title('Zero-to-next median seconds');cc=pd.DataFrame(crosses);ax[1].bar(cc.dataset_type,cc.overlap_rate*100);ax[1].set_title('Cross-stream overlap %');fig.tight_layout();fig.savefig(out/f'호텔검색_관측형합성1000명_시간간격교차스트림시각화_{stamp}.png',dpi=160);plt.close(fig)
 c.close();source.close();assert sha(a.db)==before;print(json.dumps({'status':'PASS','core':payload['core_pairs'],'timing':timing,'cross':crosses},ensure_ascii=False,indent=2))
if __name__=='__main__':main()
